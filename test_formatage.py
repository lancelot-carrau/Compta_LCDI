#!/usr/bin/env python3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Test simple pour vérifier le formatage
wb = Workbook()
ws = wb.active

# Ajouter des données de test
ws['A1'] = 'Test'
ws['B1'] = 'Shopify'
ws['C1'] = 'Statut'

ws['A2'] = 'Commande 1'
ws['B2'] = 150.0
ws['C2'] = 'COMPLET'

ws['A3'] = 'Commande 2'  
ws['B3'] = 0
ws['C3'] = 'INCOMPLET'

# Styles
header_font = Font(bold=True)
shopify_font = Font(color='FF0000', bold=True)
shopify_content_font = Font(color='FF0000')
complete_fill = PatternFill(start_color='66CC66', end_color='66CC66', fill_type='solid')
incomplete_fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')

# Appliquer les styles
ws['A1'].font = header_font
ws['B1'].font = shopify_font  # En-tête Shopify en rouge
ws['C1'].font = header_font

ws['B2'].font = shopify_content_font  # Contenu Shopify en rouge
ws['C2'].fill = complete_fill

ws['C3'].fill = incomplete_fill

wb.save('test_formatage.xlsx')
print('✅ Fichier test créé: test_formatage.xlsx')
