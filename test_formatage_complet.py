#!/usr/bin/env python3
"""
Test complet du nouveau formatage Excel avec cellules vides en rouge clair
"""

import pandas as pd
import numpy as np
import tempfile
import os

def test_formatage_complet():
    """Test complet du formatage des cellules vides"""
    
    # Simuler des données réelles avec toutes les colonnes
    test_data = {
        'Réf.WEB': ['#1001', '#1002', '#1003', '#1004', '#1005'],
        'Réf. LMB': ['LMB001', np.nan, 'LMB003', '', 'LMB005'],  # Certaines manquantes
        'Date Facture': ['2024-01-01', '2024-01-02', np.nan, '2024-01-04', ''],  # Certaines manquantes
        'Client': ['Client A', '', 'Client C', 'Client D', np.nan],  # Certaines manquantes
        'Etat': ['Payé', 'En cours', np.nan, 'Annulé', 'Payé'],  # Certaines manquantes
        'HT': [83.33, 166.67, 250.00, 0, 41.67],
        'TVA': [16.67, 33.33, 50.00, 0, 8.33],
        'TTC': [100, 200, 300, 0, 50],
        'reste': [0, 50, 100, 0, 0],
        'Shopify': [100, 150, 200, 0, 50],
        'Frais de commission': [5, 7.5, 10, 0, 2.5],
        'Virement bancaire': [0, 0, 0, 0, 0],
        'ALMA': [0, 0, 0, 0, 0],
        'Younited': [0, 0, 0, 0, 0],
        'PayPal': [95, 142.5, 190, 0, 47.5],
        'Centre de profit': ['lcdi.fr'] * 5,  # Forcé partout
        'Statut': ['COMPLET', 'INCOMPLET', 'INCOMPLET', 'INCOMPLET', 'INCOMPLET']
    }
    
    df_test = pd.DataFrame(test_data)
    
    print("📊 Données de test créées:")
    print(f"   - {len(df_test)} lignes")
    print(f"   - {len(df_test.columns)} colonnes")
    
    # Compter les cellules vides dans les colonnes importantes
    important_cols = ['Réf. LMB', 'Date Facture', 'Client', 'Etat']
    empty_count = 0
    for col in important_cols:
        empty_in_col = df_test[col].isna().sum() + (df_test[col] == '').sum()
        print(f"   - {col}: {empty_in_col} cellules vides")
        empty_count += empty_in_col
    
    print(f"   - Total: {empty_count} cellules vides à formater en rouge")
    
    # Créer fichier temporaire
    with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
        temp_path = tmp.name
    
    try:
        # Importer la fonction depuis app.py
        from app import save_with_conditional_formatting
        
        # Sauvegarder avec formatage
        result_path, is_excel = save_with_conditional_formatting(df_test, temp_path)
        
        if is_excel:
            print(f"\n✅ Fichier Excel créé: {result_path}")
            
            # Vérifier le formatage
            from openpyxl import load_workbook
            wb = load_workbook(result_path)
            ws = wb.active
            
            formatted_count = 0
            print("\n🎨 Cellules formatées en rouge clair:")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                for col_idx, cell in enumerate(row):
                    if cell.fill.start_color.index != '00000000':  # Cellule colorée
                        col_name = ws.cell(1, col_idx + 1).value
                        formatted_count += 1
                        print(f"   ✓ {col_name} ligne {row_idx} (valeur: {cell.value})")
            
            print(f"\n📈 Résumé:")
            print(f"   - Cellules vides attendues: {empty_count}")
            print(f"   - Cellules formatées: {formatted_count}")
            
            if formatted_count == empty_count:
                print("🎉 PARFAIT: Toutes les cellules vides sont formatées!")
            else:
                print("⚠️ ATTENTION: Différence entre attendu et formaté")
                
        else:
            print("❌ Erreur: Fallback vers CSV au lieu d'Excel")
            
    except Exception as e:
        print(f"❌ Erreur durant le test: {e}")
        import traceback
        traceback.print_exc()
        
    finally:
        # Nettoyer
        for ext in ['.csv', '.xlsx']:
            test_file = temp_path.replace('.csv', ext)
            if os.path.exists(test_file):
                try:
                    os.unlink(test_file)
                except:
                    pass

if __name__ == "__main__":
    test_formatage_complet()
