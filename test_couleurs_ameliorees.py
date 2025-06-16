#!/usr/bin/env python3
"""
Test du nouveau formatage avec couleurs améliorées :
- Rouge plus sombre pour cellules vides et INCOMPLET
- Vert clair pour COMPLET
"""

import pandas as pd
import numpy as np
import tempfile
import os

def test_nouveau_formatage():
    """Test du formatage coloré amélioré"""
    
    # Créer des données de test avec différents statuts
    test_data = {
        'Réf. LMB': ['LMB001', np.nan, 'LMB003', '', 'LMB005'],
        'Date Facture': ['2024-01-01', '2024-01-02', np.nan, '2024-01-04', '2024-01-05'],
        'Client': ['Client A', '', 'Client C', 'Client D', np.nan],
        'Etat': ['Payé', 'En cours', np.nan, 'Annulé', 'Payé'],
        'TTC': [100, 200, 300, 0, 50],
        'Statut': ['COMPLET', 'INCOMPLET', 'INCOMPLET', 'INCOMPLET', 'COMPLET']
    }
    
    df_test = pd.DataFrame(test_data)
    
    print("📊 Test du nouveau formatage coloré:")
    print("   🔴 Rouge plus sombre : cellules vides + INCOMPLET")
    print("   🟢 Vert clair : COMPLET")
    print(f"   📝 {len(df_test)} lignes de test")
    
    # Compter les éléments à formater
    complet_count = (df_test['Statut'] == 'COMPLET').sum()
    incomplet_count = (df_test['Statut'] == 'INCOMPLET').sum()
    
    important_cols = ['Réf. LMB', 'Date Facture', 'Client', 'Etat']
    empty_count = 0
    for col in important_cols:
        empty_in_col = df_test[col].isna().sum() + (df_test[col] == '').sum()
        empty_count += empty_in_col
    
    print(f"   🟢 COMPLET : {complet_count} cellules")
    print(f"   🔴 INCOMPLET : {incomplet_count} cellules") 
    print(f"   🔴 Cellules vides : {empty_count} cellules")
    
    # Créer fichier temporaire
    with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
        temp_path = tmp.name
    
    try:
        # Importer et utiliser la fonction de sauvegarde
        from app import save_with_conditional_formatting
        
        result_path, is_excel = save_with_conditional_formatting(df_test, temp_path)
        
        if is_excel:
            print(f"\n✅ Fichier Excel créé: {os.path.basename(result_path)}")
            
            # Analyser le formatage
            from openpyxl import load_workbook
            wb = load_workbook(result_path)
            ws = wb.active
            
            print("\n🎨 Analyse du formatage:")
              # Compter les différents formatages
            rouge_count = 0
            vert_count = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                for col_idx, cell in enumerate(row):
                    if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        col_name = ws.cell(1, col_idx + 1).value
                        
                        # Récupérer la couleur (peut être RGB ou index)
                        color_value = None
                        if hasattr(cell.fill.start_color, 'rgb'):
                            color_value = cell.fill.start_color.rgb
                        elif hasattr(cell.fill.start_color, 'index'):
                            color_value = cell.fill.start_color.index
                        
                        # Identifier les couleurs par leur valeur RGB/hex
                        if color_value and str(color_value).upper() in ['FFB3B3', 'FFFFB3B3']:  # Rouge plus sombre
                            rouge_count += 1
                            if col_name == 'Statut':
                                print(f"   🔴 {col_name} ligne {row_idx}: {cell.value}")
                            else:
                                print(f"   🔴 {col_name} ligne {row_idx}: cellule vide")
                        elif color_value and str(color_value).upper() in ['B3FFB3', 'FFB3FFB3']:  # Vert clair
                            vert_count += 1
                            print(f"   🟢 {col_name} ligne {row_idx}: {cell.value}")
                        elif color_value != '00000000':  # Toute couleur non-blanche
                            print(f"   🎨 {col_name} ligne {row_idx}: couleur {color_value} (valeur: {cell.value})")
                            # Compter comme rouge ou vert selon le contexte
                            if col_name == 'Statut':
                                if cell.value == 'COMPLET':
                                    vert_count += 1
                                elif cell.value == 'INCOMPLET':
                                    rouge_count += 1
                            else:
                                rouge_count += 1
            
            print(f"\n📊 Résultats:")
            print(f"   🔴 Cellules rouges: {rouge_count}")
            print(f"   🟢 Cellules vertes: {vert_count}")
            
            expected_rouge = empty_count + incomplet_count
            if rouge_count == expected_rouge and vert_count == complet_count:
                print("🎉 PARFAIT! Formatage correct:")
                print(f"   ✓ Rouge: {rouge_count}/{expected_rouge} (cellules vides + INCOMPLET)")
                print(f"   ✓ Vert: {vert_count}/{complet_count} (COMPLET)")
            else:
                print("⚠️ Différences détectées:")
                print(f"   Attendu rouge: {expected_rouge}, obtenu: {rouge_count}")
                print(f"   Attendu vert: {complet_count}, obtenu: {vert_count}")
                
        else:
            print("❌ Fallback vers CSV - openpyxl non disponible")
            
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        
    finally:
        # Nettoyer
        for ext in ['.csv', '.xlsx']:
            test_file = temp_path.replace('.csv', ext)
            if os.path.exists(test_file):
                try:
                    os.unlink(test_file)
                except:
                    pass

if __name__ == "__main__":
    test_nouveau_formatage()
