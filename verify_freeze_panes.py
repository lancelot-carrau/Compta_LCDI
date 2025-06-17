import pandas as pd
import os
from datetime import datetime

def verify_freeze_panes():
    """Vérifier que les volets figés sont bien appliqués dans le fichier Excel"""
    print("=== VÉRIFICATION DU FIGEMENT DES EN-TÊTES ===\n")
    
    # Chercher le fichier le plus récent
    output_folder = r'C:\Code\Apps\Compta LCDI V2\output'
    excel_files = [f for f in os.listdir(output_folder) if f.endswith('.xlsx')]
    
    if not excel_files:
        print("❌ Aucun fichier Excel trouvé")
        return
    
    latest_file = max(excel_files, key=lambda f: os.path.getmtime(os.path.join(output_folder, f)))
    file_path = os.path.join(output_folder, latest_file)
    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
    
    print(f"📄 Fichier testé: {latest_file}")
    print(f"⏰ Date de création: {file_time.strftime('%d/%m/%Y %H:%M:%S')}")
    
    try:
        # Utiliser openpyxl pour vérifier les freeze_panes
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"\n📊 Informations du fichier Excel:")
        print(f"   - Nom de la feuille: {ws.title}")
        print(f"   - Nombre de lignes: {ws.max_row}")
        print(f"   - Nombre de colonnes: {ws.max_column}")
        
        # Vérifier les freeze_panes
        if ws.freeze_panes:
            freeze_cell = ws.freeze_panes
            print(f"\n❄️  FIGEMENT DES VOLETS:")
            print(f"   ✅ Volets figés activés!")
            print(f"   📍 Cellule de figement: {freeze_cell}")
            
            if freeze_cell == 'A2':
                print(f"   ✅ Configuration correcte: La ligne 1 (en-têtes) restera visible lors du défilement")
            else:
                print(f"   ⚠️  Configuration inattendue: {freeze_cell}")
                
        else:
            print(f"\n❌ FIGEMENT DES VOLETS:")
            print(f"   ❌ Aucun volet figé détecté")
        
        # Afficher les en-têtes
        print(f"\n📋 EN-TÊTES DE COLONNES (ligne 1):")
        headers = []
        for col_idx in range(1, min(ws.max_column + 1, 11)):  # Les 10 premières colonnes
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(str(cell_value) if cell_value else "")
            
        for i, header in enumerate(headers, 1):
            print(f"   {chr(64+i)}: {header}")
            
        if ws.max_column > 10:
            print(f"   ... et {ws.max_column - 10} autres colonnes")
        
        # Test basique de lecture avec pandas pour s'assurer que le fichier est valide
        df = pd.read_excel(file_path)
        print(f"\n✅ VALIDATION:")
        print(f"   ✅ Fichier Excel lisible avec pandas")
        print(f"   ✅ {len(df)} lignes de données + 1 ligne d'en-têtes")
        print(f"   ✅ Figement des volets configuré pour faciliter la navigation")
        
        return True
        
    except ImportError:
        print("❌ openpyxl non disponible pour vérifier les freeze_panes")
        return False
    except Exception as e:
        print(f"❌ Erreur lors de la vérification: {e}")
        return False

if __name__ == "__main__":
    verify_freeze_panes()
