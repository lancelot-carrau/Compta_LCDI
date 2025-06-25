#!/usr/bin/env python3
"""
Script pour diagnostiquer la structure du fichier Excel
"""
import pandas as pd
import openpyxl

# Fichier spécifique à analyser
excel_file = "output/Compta_LCDI_Amazon_23_06_2025_14_22_06.xlsx"

print(f"📄 Diagnostic du fichier: {excel_file}")
print("=" * 60)

try:
    # Lire avec openpyxl pour voir la structure brute
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    print(f"📊 Feuille active: {sheet.title}")
    print(f"📏 Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes")
    print()
    
    # Afficher les premières lignes brutes
    print("📋 Premières lignes (brutes):")
    for row_num in range(1, min(6, sheet.max_row + 1)):
        row_data = []
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"   Ligne {row_num}: {row_data}")
    
    print()
    
    # Essayer de lire avec pandas en sautant les en-têtes
    print("📊 Tentative de lecture pandas:")
    df = pd.read_excel(excel_file, header=None)
    print(f"   Forme du DataFrame: {df.shape}")
    print(f"   Premières lignes:")
    print(df.head())
    
except Exception as e:
    print(f"❌ Erreur: {e}")
