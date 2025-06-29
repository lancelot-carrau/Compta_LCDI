from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
import os
from datetime import datetime
import tempfile
from werkzeug.utils import secure_filename
import io
import chardet
import re
import logging
import traceback
import sys

# Configuration du logging avec sortie console forcée
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Middleware pour logger TOUTES les requêtes
@app.before_request
def log_request_info():
    print(f"\n{'='*60}")
    print(f"NOUVELLE REQUÊTE REÇUE:")
    print(f"Method: {request.method}")
    print(f"URL: {request.url}")
    print(f"Path: {request.path}")
    print(f"Endpoint: {request.endpoint}")
    print(f"Headers: {dict(request.headers)}")
    print(f"Form data: {dict(request.form) if request.form else 'Aucune'}")
    print(f"Files: {list(request.files.keys()) if request.files else 'Aucun'}")
    print(f"Args: {dict(request.args) if request.args else 'Aucun'}")
    print(f"{'='*60}\n")
    
    logger.info(f"REQUEST: {request.method} {request.path}")
    logger.debug(f"Headers: {dict(request.headers)}")
    logger.debug(f"Form: {dict(request.form) if request.form else 'Empty'}")
    logger.debug(f"Files: {list(request.files.keys()) if request.files else 'Empty'}")

@app.after_request
def log_response_info(response):
    print(f"\n{'='*60}")
    print(f"RÉPONSE ENVOYÉE:")
    print(f"Status: {response.status_code}")
    print(f"Headers: {dict(response.headers)}")
    print(f"{'='*60}\n")
    
    logger.info(f"RESPONSE: {response.status_code}")
    return response

# Gestionnaire d'erreurs global
@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Exception non gérée: {str(e)}")
    logger.error(f"Traceback complet: {traceback.format_exc()}")
    flash(f"Erreur inattendue: {str(e)}", 'error')
    return render_template('index.html'), 500

@app.errorhandler(405)
def method_not_allowed(e):
    print(f"\n{'!'*80}")
    print(f"ERREUR 405 METHOD NOT ALLOWED DÉTECTÉE!")
    print(f"Method: {request.method}")
    print(f"URL: {request.url}")
    print(f"Path: {request.path}")
    print(f"Endpoint: {request.endpoint}")
    print(f"User-Agent: {request.headers.get('User-Agent', 'Unknown')}")
    print(f"Referer: {request.headers.get('Referer', 'Unknown')}")
    print(f"Content-Type: {request.headers.get('Content-Type', 'Unknown')}")
    print(f"Headers complets: {dict(request.headers)}")
    print(f"Form data: {dict(request.form) if request.form else 'Aucune'}")
    print(f"Files: {list(request.files.keys()) if request.files else 'Aucun'}")
    print(f"Args: {dict(request.args) if request.args else 'Aucun'}")
    print(f"{'!'*80}\n")
    
    logger.error(f"Erreur 405 Method Not Allowed: {request.method} {request.url}")
    logger.error(f"Headers: {dict(request.headers)}")
    logger.error(f"Form data: {request.form}")
    logger.error(f"Files: {request.files}")
    
    # Retourner une réponse JSON pour les requêtes AJAX
    if request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
        return {'error': 'Method not allowed', 'method': request.method, 'path': request.path}, 405
    
    flash("Méthode HTTP non autorisée pour cette route", 'error')
    return render_template('index.html'), 405

@app.errorhandler(413)
def too_large(e):
    logger.error(f"Fichier trop volumineux: {request.url}")
    flash("Le fichier est trop volumineux (max 16MB)", 'error')
    return render_template('index.html'), 413

# Configuration des dossiers
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'csv'}

# Créer les dossiers s'ils n'existent pas
for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

def allowed_file(filename):
    """Vérifie si l'extension du fichier est autorisée"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def detect_encoding(file_path):
    """Détecte automatiquement l'encodage d'un fichier"""
    try:
        # Lire les premiers bytes du fichier pour détecter l'encodage
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)  # Lire les premiers 10KB
        
        # Utiliser chardet pour détecter l'encodage
        result = chardet.detect(raw_data)
        encoding = result['encoding']
        confidence = result['confidence']
        
        print(f"Encodage détecté pour {file_path}: {encoding} (confiance: {confidence:.2f})")
        
        # Liste des encodages fallback en ordre de préférence
        fallback_encodings = ['utf-8', 'utf-8-sig', 'windows-1252', 'iso-8859-1', 'cp1252', 'latin-1']
        
        # Si la confiance est faible ou l'encodage n'est pas détecté, essayer les fallbacks
        if not encoding or confidence < 0.7:
            print(f"Confiance faible ({confidence:.2f}), essai des encodages fallback...")
            for fallback in fallback_encodings:
                try:
                    with open(file_path, 'r', encoding=fallback) as f:
                        f.read(1000)  # Essayer de lire le début du fichier
                    print(f"Encodage fallback réussi: {fallback}")
                    return fallback
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            # Si tous les fallbacks échouent, utiliser latin-1 (qui peut lire n'importe quoi)
            print("Tous les encodages ont échoué, utilisation de latin-1")
            return 'latin-1'
        
        return encoding
        
    except Exception as e:
        print(f"Erreur lors de la détection d'encodage: {e}")
        # En cas d'erreur, essayer les encodages les plus courants
        for encoding in ['utf-8', 'windows-1252', 'iso-8859-1', 'latin-1']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1000)
                print(f"Encodage de secours utilisé: {encoding}")
                return encoding
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        # Dernier recours
        return 'latin-1'

def safe_read_csv(file_path, separator=','):
    """Lit un fichier CSV avec détection automatique de l'encodage"""
    encoding = detect_encoding(file_path)
    
    # Essayer de lire le fichier avec l'encodage détecté
    try:
        df = pd.read_csv(file_path, sep=separator, encoding=encoding)
        print(f"Fichier lu avec succès avec l'encodage {encoding}")
        return df
    except (UnicodeDecodeError, UnicodeError) as e:
        print(f"Erreur avec l'encodage {encoding}: {e}")
        
        # Essayer d'autres encodages en cas d'échec
        fallback_encodings = ['utf-8', 'utf-8-sig', 'windows-1252', 'iso-8859-1', 'cp1252', 'latin-1']
        
        for fallback in fallback_encodings:
            if fallback == encoding:  # Skip l’encodage déjà essayé
                continue
            try:
                df = pd.read_csv(file_path, sep=separator, encoding=fallback)
                print(f"Fichier lu avec succès avec l'encodage fallback {fallback}")
                return df
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        # Si tout échoue, lever l'erreur originale
        raise e

def normalize_column_names(df, expected_columns, file_type=""):
    """Normalise les noms de colonnes en cherchant des correspondances approximatives"""
    print(f"Colonnes disponibles dans {file_type}: {list(df.columns)}")    # Dictionnaire de mapping pour les variantes de noms de colonnes
    column_mappings = {
        # Fichier des commandes
        'Name': ['Name', 'name', 'ORDER', 'Order', 'order', 'Nom', 'nom', 'Commande', 'commande', 'Id', 'ID', 'id', '#', 'Order ID', 'Order Id'],
        'Fulfilled at': ['Fulfilled at', 'fulfilled at', 'Date', 'date', 'Date commande', 'Date_commande'],
        'Billing name': ['Billing name', 'billing name', 'Client', 'client', 'Nom client', 'Nom_client', 'Billing Name', 'billing Name'],
        'Financial Status': ['Financial Status', 'financial status', 'Status', 'status', 'Statut', 'statut'],
        'Tax 1 Value': ['Tax 1 Value', 'tax 1 value', 'TVA', 'tva', 'Tax', 'tax', 'Taxe'],
        'Outstanding Balance': ['Outstanding Balance', 'outstanding balance', 'Balance', 'balance', 'Solde', 'solde'],
        'Payment Method': ['Payment Method', 'payment method', 'Method', 'method', 'Méthode', 'méthode'],
        # IMPORTANT: Colonnes pour le nouveau calcul HT/TVA/TTC
        'Total': ['Total', 'total'],  # TTC - UNIQUEMENT la vraie colonne Total !
        'Taxes': ['Taxes', 'taxes'],  # TVA - UNIQUEMENT la vraie colonne Taxes !
        
        # Fichier des transactions
        'Order': ['Order', 'order', 'Name', 'name', 'Commande', 'commande', 'Id', 'ID', 'id', 'Order ID', 'Order Id'],
        'Presentment Amount': ['Presentment Amount', 'presentment amount', 'Amount', 'amount', 'Montant', 'montant'],
        'Fee': ['Fee', 'fee', 'Frais', 'frais', 'Commission', 'commission'],
        'Net': ['Net', 'net', 'Net Amount', 'net amount', 'Montant net', 'montant net'],
          # Fichier journal
        'Piece': ['Piece', 'piece', 'Pièce', 'pièce', 'Reference', 'reference', 'Ref', 'ref', 'Order', 'order', 'Commande', 'Id', 'ID', 'id', 'Référence externe', 'référence externe', 'Reference externe', 'reference externe', 'Externe', 'externe'],
        'Référence LMB': ['Référence LMB', 'référence lmb', 'Reference LMB', 'reference lmb', 'LMB', 'lmb', 'Ref LMB', 'ref lmb']
    }
    
    # Créer un mapping des colonnes actuelles vers les noms standardisés
    column_mapping = {}
    missing_columns = []
    
    for expected_col in expected_columns:
        found = False
        if expected_col in column_mappings:
            for variant in column_mappings[expected_col]:
                if variant in df.columns:
                    column_mapping[variant] = expected_col
                    print(f"✓ Colonne trouvée: '{variant}' -> '{expected_col}'")
                    found = True
                    break
        
        if not found:
            missing_columns.append(expected_col)
      # Afficher les colonnes manquantes
    if missing_columns:
        print(f"⚠️ Colonnes manquantes dans {file_type}: {missing_columns}")
        print("Colonnes disponibles:", list(df.columns))
        
        # Essayer une correspondance approximative pour les colonnes manquantes
        for missing_col in missing_columns:
            found_match = False
            
            # Vérification exacte des mots-clés
            keywords_mapping = {
                'Name': ['id', 'order', 'commande', 'numero', 'number'],
                'Fulfilled at': ['date', 'created', 'fulfill', 'livr'],
                'Billing name': ['billing', 'client', 'nom', 'name'],
                'Financial Status': ['status', 'statut', 'financial', 'etat'],
                'Tax 1 Value': ['tax', 'tva', 'taxe', 'impot'],                'Outstanding Balance': ['balance', 'solde', 'outstanding', 'restant'],
                'Payment Method': ['payment', 'method', 'paiement', 'methode'],
                'Order': ['order', 'id', 'commande', 'numero', 'name'],
                'Presentment Amount': ['amount', 'montant', 'presentment'],
                'Fee': ['fee', 'frais', 'commission'],
                'Net': ['net', 'montant', 'amount'],
                # IMPORTANT: Pas de mapping approximatif pour Total et Taxes !
                # 'Total': sera géré par mapping exact uniquement
                # 'Taxes': sera géré par mapping exact uniquement
                'Piece': ['piece', 'reference', 'ref', 'order', 'id', 'commande', 'externe', 'external'],
                'Référence LMB': ['lmb', 'reference', 'ref']
            }
            
            if missing_col in keywords_mapping:
                keywords = keywords_mapping[missing_col]
                for col in df.columns:
                    col_lower = col.lower()
                    # Vérifier si un des mots-clés est dans le nom de la colonne
                    if any(keyword in col_lower for keyword in keywords):
                        print(f"🔍 Correspondance trouvée: '{col}' pour '{missing_col}' (mot-clé détecté)")
                        column_mapping[col] = missing_col
                        found_match = True
                        break
              # Si toujours pas trouvé, essayer une correspondance plus flexible
            if not found_match:
                for col in df.columns:
                    # PROTECTION: Empêcher que "Subtotal" soit mappé vers "Total"
                    if missing_col == 'Total' and 'subtotal' in col.lower():
                        print(f"❌ REJETÉ: '{col}' ne peut pas être mappé vers 'Total' (doit être la vraie colonne Total)")
                        continue
                    
                    # PROTECTION: Empêcher que d'autres colonnes soient mappées vers "Taxes"
                    if missing_col == 'Taxes' and col.lower() not in ['taxes', 'tax']:
                        continue
                    
                    # Vérification de similarité (contient une partie du nom)
                    if any(part.lower() in col.lower() for part in missing_col.lower().split() if len(part) > 2):
                        print(f"🔍 Correspondance approximative: '{col}' pour '{missing_col}'")
                        # Demander confirmation via un message de debug
                        print(f"   -> Voulez-vous utiliser '{col}' pour '{missing_col}' ? (automatiquement accepté)")
                        column_mapping[col] = missing_col
                        found_match = True
                        break
    
    # Renommer les colonnes
    if column_mapping:
        df = df.rename(columns=column_mapping)
        print(f"✓ Colonnes renommées: {column_mapping}")
    
    return df

def validate_required_columns(df, required_columns, file_type=""):
    """Valide que toutes les colonnes requises sont présentes"""
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        available_cols = list(df.columns)
        error_msg = f"""
        Colonnes manquantes dans {file_type}: {missing_columns}
        
        Colonnes disponibles dans le fichier: {available_cols}
        
        Vérifiez que votre fichier {file_type} contient bien les colonnes requises.
        """
        raise ValueError(error_msg)
    
    print(f"✓ Toutes les colonnes requises sont présentes dans {file_type}")
    return True

def clean_text_data(df, text_columns):
    """Nettoie les données texte en supprimant les espaces superflus"""
    for col in text_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    return df

def format_date_to_french(date_str):
    """Convertit une date en format français jj/mm/aaaa"""
    try:
        if pd.isna(date_str) or str(date_str).lower() in ['nan', 'none', '']:
            return ''
        
        # Essayer différents formats de date d'entrée
        date_formats = [
            '%Y-%m-%d',           # 2024-12-25
            '%d/%m/%Y',           # 25/12/2024
            '%m/%d/%Y',           # 12/25/2024
            '%Y-%m-%d %H:%M:%S',  # 2024-12-25 10:30:00
            '%d-%m-%Y',           # 25-12-2024
            '%Y/%m/%d'            # 2024/12/25
        ]
        
        for fmt in date_formats:
            try:
                dt = pd.to_datetime(date_str, format=fmt)
                return dt.strftime('%d/%m/%Y')
            except:
                continue
                
        # Si aucun format ne marche, essayer la conversion automatique pandas
        try:
            dt = pd.to_datetime(date_str)
            return dt.strftime('%d/%m/%Y')
        except:
            return str(date_str)
            
    except Exception as e:
        print(f"Erreur lors du formatage de la date '{date_str}': {e}")
        return str(date_str) if date_str else ''

def categorize_payment_method(payment_method_orders, payment_method_transactions, ttc_value, fallback_amount=None):
    """
    Catégorise les méthodes de paiement et retourne les montants par catégorie
    NOUVELLE LOGIQUE:
    - Les paiements par carte (Shopify Payments, credit_card, etc.) vont dans "Carte bancaire"
    - Utilise prioritairement Payment Method Name du fichier transactions pour PayPal
    - Laisse les cellules vides pour les méthodes non reconnues
    """
    # Initialiser toutes les catégories à 0
    result = {'Virement bancaire': 0, 'Carte bancaire': 0, 'ALMA': 0, 'Younited': 0, 'PayPal': 0}
    
    # Utiliser le montant principal, sinon le fallback
    amount_to_use = ttc_value
    if pd.isna(ttc_value) and fallback_amount is not None and not pd.isna(fallback_amount):
        amount_to_use = fallback_amount
        print(f"DEBUG: Utilisation fallback amount {fallback_amount} pour méthode orders='{payment_method_orders}', transactions='{payment_method_transactions}'")
    
    # Si aucun montant valide, retourner 0 partout
    if pd.isna(amount_to_use):
        print(f"DEBUG: Aucun montant valide pour méthodes orders='{payment_method_orders}', transactions='{payment_method_transactions}', retour 0")
        return result
    
    ttc_amount = float(amount_to_use)    # Préparer les chaînes pour la comparaison
    payment_orders_str = str(payment_method_orders).lower() if not pd.isna(payment_method_orders) else ""
    payment_transactions_str = str(payment_method_transactions).lower() if not pd.isna(payment_method_transactions) else ""
    
    print(f"DEBUG: Analyse paiement - Orders: '{payment_orders_str}', Transactions: '{payment_transactions_str}', Montant: {ttc_amount}")
    
    # DEBUG SPÉCIAL pour les commandes PayPal problématiques
    if any(ref in str(payment_orders_str + payment_transactions_str) for ref in ['1041', '1037', '1040', '1042']):
        print(f"🔍 DEBUG COMMANDE SPÉCIALE: Orders='{payment_method_orders}', Transactions='{payment_method_transactions}', TTC={ttc_amount}")
    
    # PRIORITÉ 1: Vérifier PayPal dans les transactions (plus précis)
    # Détection élargie : paypal, pay pal, pp, etc.
    if ('paypal' in payment_transactions_str or 'pay pal' in payment_transactions_str or 
        'pay-pal' in payment_transactions_str or payment_transactions_str == 'pp'):
        result['PayPal'] = ttc_amount
        print(f"DEBUG: PayPal détecté dans transactions -> PayPal: {ttc_amount}")
    # AUSSI: Vérifier PayPal dans les commandes (fallback)
    elif ('paypal' in payment_orders_str or 'pay pal' in payment_orders_str or 
          'pay-pal' in payment_orders_str or payment_orders_str == 'pp'):
        result['PayPal'] = ttc_amount
        print(f"DEBUG: PayPal détecté dans commandes -> PayPal: {ttc_amount}")
    # PRIORITÉ 2: Alma et Younited
    elif 'alma' in payment_orders_str or 'alma' in payment_transactions_str:
        result['ALMA'] = ttc_amount
        print(f"DEBUG: ALMA détecté -> ALMA: {ttc_amount}")
    elif 'younited' in payment_orders_str or 'younited' in payment_transactions_str:
        result['Younited'] = ttc_amount
        print(f"DEBUG: Younited détecté -> Younited: {ttc_amount}")
    # PRIORITÉ 3: Vrais virements bancaires uniquement
    elif ('virement' in payment_orders_str or 'wire' in payment_orders_str or 'bank' in payment_orders_str or
          'custom' in payment_orders_str):  # Custom = souvent virement bancaire
        result['Virement bancaire'] = ttc_amount
        print(f"DEBUG: Virement bancaire détecté -> Virement bancaire: {ttc_amount}")    # PRIORITÉ 4: Paiements par carte bancaire
    elif ('shopify payments' in payment_orders_str or 'shopify payment' in payment_orders_str or
          'credit_card' in payment_orders_str or 'credit card' in payment_orders_str or
          'carte' in payment_transactions_str or 'card' in payment_transactions_str):
        # Paiements par carte: vont dans la colonne "Carte bancaire"
        result['Carte bancaire'] = ttc_amount
        print(f"DEBUG: Paiement par carte détecté -> Carte bancaire: {ttc_amount}")
    else:
        # Méthode non reconnue, laisser les cellules vides pour traitement manuel
        print(f"DEBUG: Méthode de paiement non reconnue: orders='{payment_orders_str}', transactions='{payment_transactions_str}' -> cellules vides")
        # Toutes les catégories restent à 0 (cellules vides)
    
    return result

def calculate_corrected_amounts(df_merged_final):
    """
    Calcule les montants HT, TVA, TTC avec logique stricte et fallback conditionnel.
    PRIORITÉ 1: Utilise strictement les colonnes du Journal ("Montant du document TTC", "Montant du document HT")
    PRIORITÉ 2: Si pas de données Journal OU de transactions, laisse la cellule vide (NaN) pour formatage conditionnel rouge
    FALLBACK CONDITIONNEL: Si TTC, HT, TVA ET Shopify sont TOUS vides sur une ligne, 
                           alors utiliser "Total" et "Taxes" du fichier commandes UNIQUEMENT pour ces lignes
    """
    # Debug: afficher les colonnes disponibles
    print(f"DEBUG: Colonnes disponibles: {list(df_merged_final.columns)}")
    
    # Vérifier s'il y a des doublons
    column_counts = df_merged_final.columns.value_counts()
    duplicates = column_counts[column_counts > 1]
    if not duplicates.empty:
        print(f"DEBUG: Colonnes dupliquées trouvées: {duplicates.to_dict()}")
        # Supprimer les doublons en gardant la première occurrence
        df_merged_final = df_merged_final.loc[:, ~df_merged_final.columns.duplicated()]
        print(f"DEBUG: Colonnes après suppression des doublons: {list(df_merged_final.columns)}")
    
    # Initialiser toutes les séries avec NaN (cellules vides) et le bon index
    ttc_amounts = pd.Series([None] * len(df_merged_final), dtype=float, index=df_merged_final.index)
    ht_amounts = pd.Series([None] * len(df_merged_final), dtype=float, index=df_merged_final.index)
    tva_amounts = pd.Series([None] * len(df_merged_final), dtype=float, index=df_merged_final.index)
    
    # ÉTAPE 1: Traiter les montants du Journal (strictement prioritaires)
    journal_ttc_available = 'Montant du document TTC' in df_merged_final.columns
    journal_ht_available = 'Montant du document HT' in df_merged_final.columns
    
    if journal_ttc_available:
        print("DEBUG: Traitement des montants TTC du Journal (priorité absolue)")
        # Convertir les montants français (virgule) en format numérique
        ttc_col = df_merged_final['Montant du document TTC'].astype(str).str.replace(',', '.').str.replace(' ', '')
        ttc_amounts_journal = pd.to_numeric(ttc_col, errors='coerce')
        
        # Appliquer les montants du journal là où ils existent
        mask_journal_ttc = ttc_amounts_journal.notna()
        ttc_amounts.loc[mask_journal_ttc] = ttc_amounts_journal.loc[mask_journal_ttc]
        print(f"DEBUG: {mask_journal_ttc.sum()} montants TTC récupérés du Journal")
        
    if journal_ht_available:
        print("DEBUG: Traitement des montants HT du Journal (priorité absolue)")
        # Convertir les montants français (virgule) en format numérique
        ht_col = df_merged_final['Montant du document HT'].astype(str).str.replace(',', '.').str.replace(' ', '')
        ht_amounts_journal = pd.to_numeric(ht_col, errors='coerce')
        
        # Appliquer les montants HT du journal là où ils existent
        mask_journal_ht = ht_amounts_journal.notna()
        ht_amounts.loc[mask_journal_ht] = ht_amounts_journal.loc[mask_journal_ht]
        print(f"DEBUG: {mask_journal_ht.sum()} montants HT récupérés du Journal")
        
        # Calculer TVA = TTC - HT (seulement là où on a les deux du journal)
        mask_both_journal = ttc_amounts.notna() & ht_amounts.notna()
        tva_amounts.loc=mask_both_journal] = ttc_amounts.loc[mask_both_journal] - ht_amounts.loc[mask_both_journal]
        print(f"DEBUG: {mask_both_journal.sum()} montants TVA calculés depuis Journal (TTC - HT)")
      # ÉTAPE 2: Appliquer le fallback conditionnel
    # Condition: TTC, HT, TVA sont TOUS vides (peu importe le statut de Shopify)
    print("DEBUG: Application du fallback conditionnel...")
    
    # Identifier les lignes où les montants principaux sont vides (TTC, HT, TVA)
    mask_amounts_empty = (
        ttc_amounts.isna() & 
        ht_amounts.isna() & 
        tva_amounts.isna()
    )
    
    lines_for_fallback = mask_amounts_empty.sum()
    print(f"DEBUG: {lines_for_fallback} lignes éligibles au fallback (TTC, HT, TVA tous vides, peu importe Shopify)")
    
    # Appliquer le fallback uniquement pour ces lignes
    if lines_for_fallback > 0 and 'Total' in df_merged_final.columns:
        print("DEBUG: Application du fallback depuis les commandes (Total et Taxes)")
        
        # Récupérer les montants des commandes
        total_from_orders = pd.to_numeric(df_merged_final['Total'], errors='coerce')
        taxes_from_orders = pd.Series([None] * len(df_merged_final), dtype=float, index=df_merged_final.index)
        
        if 'Taxes' in df_merged_final.columns:
            taxes_from_orders = pd.to_numeric(df_merged_final['Taxes'], errors='coerce')
          # Appliquer le fallback UNIQUEMENT aux lignes éligibles
        mask_fallback_ttc = mask_amounts_empty & total_from_orders.notna()
        mask_fallback_tva = mask_amounts_empty & taxes_from_orders.notna()
        
        ttc_amounts.loc[mask_fallback_ttc] = total_from_orders.loc[mask_fallback_ttc]
        tva_amounts.loc[mask_fallback_tva] = taxes_from_orders.loc[mask_fallback_tva]
        
        # Calculer HT = TTC - TVA pour les lignes de fallback
        mask_fallback_ht = mask_amounts_empty & ttc_amounts.notna() & tva_amounts.notna()
        ht_amounts.loc[mask_fallback_ht] = ttc_amounts.loc[mask_fallback_ht] - tva_amounts.loc[mask_fallback_ht]
        
        print(f"DEBUG: Fallback appliqué - TTC: {mask_fallback_ttc.sum()}, TVA: {mask_fallback_tva.sum()}, HT: {mask_fallback_ht.sum()}")
    
    # Statistiques finales
    ttc_filled = ttc_amounts.notna().sum()
    ht_filled = ht_amounts.notna().sum()
    tva_filled = tva_amounts.notna().sum()
    n_rows = len(df_merged_final)
    
    print(f"DEBUG: RÉSULTAT FINAL - Cellules remplies - TTC: {ttc_filled}/{n_rows}, HT: {ht_filled}/{n_rows}, TVA: {tva_filled}/{n_rows}")
    print(f"DEBUG: Cellules vides (formatage rouge) - TTC: {n_rows - ttc_filled}, HT: {n_rows - ht_filled}, TVA: {n_rows - tva_filled}")
    print(f"DEBUG: Échantillon TTC: {ttc_amounts.head().tolist()}")
    print(f"DEBUG: Échantillon HT: {ht_amounts.head().tolist()}")
    print(f"DEBUG: Échantillon TVA: {tva_amounts.head().tolist()}")
    
    return {
        'HT': ht_amounts,
        'TVA': tva_amounts,
        'TTC': ttc_amounts
    }

def calculate_invoice_dates(df_merged_final):
    """
    Calcule les dates de facture avec logique de priorité.
    PRIORITÉ 1: Utilise "Date du document" du Journal si disponible
    PRIORITÉ 2: Utilise "Fulfilled at" des commandes (date de livraison/expédition)
    """
    print(f"DEBUG: Calcul des dates de facture...")
    
    # Vérifier les colonnes disponibles
    journal_date_available = 'Date du document' in df_merged_final.columns
    fulfilled_date_available = 'Fulfilled at' in df_merged_final.columns
    
    # Initialiser la série des dates
    invoice_dates = pd.Series([None] * len(df_merged_final))
    
    if journal_date_available:
        print("DEBUG: Utilisation prioritaire des dates du Journal")
        # Convertir les dates du journal 
        journal_dates = df_merged_final['Date du document'].copy()
        
        # Pour les lignes avec données Journal, utiliser la date du journal
        mask_journal_available = journal_dates.notna() & (journal_dates != '') & (journal_dates != 'nan')
        if mask_journal_available.any():
            # Convertir format DD/MM/YYYY HH:MM:SS vers DD/MM/YYYY
            for idx in journal_dates[mask_journal_available].index:
                date_str = str(journal_dates[idx])
                if ' ' in date_str:
                    # Enlever la partie heure
                    date_part = date_str.split(' ')[0]
                    invoice_dates[idx] = date_part
                else:
                    invoice_dates[idx] = date_str
            print(f"DEBUG: {mask_journal_available.sum()} dates récupérées du Journal")
    
    if fulfilled_date_available:
        print("DEBUG: Utilisation fallback des dates Fulfilled at")
        # Pour les lignes sans date Journal, utiliser Fulfilled at
        fulfilled_dates = df_merged_final['Fulfilled at'].copy()
        
        # Masque pour les lignes qui n'ont pas encore de date
        mask_need_fulfilled = invoice_dates.isna() & fulfilled_dates.notna() & (fulfilled_dates != '') & (fulfilled_dates != 'nan')
        
        if mask_need_fulfilled.any():
            # Convertir format ISO vers format MM/DD/YYYY
            for idx in fulfilled_dates[mask_need_fulfilled].index:
                date_str = str(fulfilled_dates[idx])
                if 'T' in date_str or '+' in date_str:
                    # Format: 2025-05-19 11:11:57 +0200 ou 2025-05-19T11:11:57+02:00
                    try:
                        # Enlever timezone et heure
                        if '+' in date_str:
                            date_str = date_str.split('+')[0]
                        if 'T' in date_str:
                            date_str = date_str.split('T')[0]
                        elif ' ' in date_str:
                            date_str = date_str.split(' ')[0]
                          # Convertir YYYY-MM-DD vers DD/MM/YYYY (format français)
                        from datetime import datetime
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        invoice_dates[idx] = date_obj.strftime('%d/%m/%Y')
                    except:
                        # En cas d'erreur, garder la date originale
                        invoice_dates[idx] = date_str
                else:
                    invoice_dates[idx] = date_str
            print(f"DEBUG: {mask_need_fulfilled.sum()} dates récupérées de Fulfilled at")
    
    # Compter les résultats
    dates_found = invoice_dates.notna().sum()
    print(f"DEBUG: Total dates de facture trouvées: {dates_found}/{len(df_merged_final)}")
    
    # Échantillon de résultats
    sample_dates = invoice_dates[invoice_dates.notna()].head(5)
    print(f"DEBUG: Échantillon dates: {sample_dates.tolist()}")
    
    return invoice_dates

def generate_consolidated_billing_table(orders_file, transactions_file, journal_file):
    """
    Fonction principale pour générer le tableau de facturation consolidé
    """
    try:
        print("=== DÉBUT DU TRAITEMENT ===")        # ÉTAPE 1: Chargement des fichiers CSV
        print("1. Chargement des fichiers CSV...")
        
        # Chargement du fichier des commandes (séparateur virgule)
        df_orders = safe_read_csv(orders_file, separator=',')
        print(f"   - Commandes chargées: {len(df_orders)} lignes")
          # Chargement du fichier des transactions (séparateur virgule)  
        df_transactions = safe_read_csv(transactions_file, separator=',')
        print(f"   - Transactions chargées: {len(df_transactions)} lignes")
        
        # Chargement du fichier journal (séparateur point-virgule)
        df_journal = safe_read_csv(journal_file, separator=';')
        print(f"   - Journal chargé: {len(df_journal)} lignes")
          # Vérification et normalisation des colonnes requises
        required_orders_cols = ['Name', 'Fulfilled at', 'Billing name', 'Financial Status', 
                               'Tax 1 Value', 'Outstanding Balance', 'Payment Method', 'Total', 'Taxes']
        required_transactions_cols = ['Order', 'Presentment Amount', 'Fee', 'Net', 'Payment Method Name']
        required_journal_cols = ['Piece', 'Référence LMB']
        
        print("\n2. Vérification et normalisation des colonnes...")
        
        # Normaliser les colonnes
        df_orders = normalize_column_names(df_orders, required_orders_cols, "fichier des commandes")
        df_transactions = normalize_column_names(df_transactions, required_transactions_cols, "fichier des transactions")
        df_journal = normalize_column_names(df_journal, required_journal_cols, "fichier journal")
          # Valider que toutes les colonnes requises sont présentes
        validate_required_columns(df_orders, required_orders_cols, "fichier des commandes")
        validate_required_columns(df_transactions, required_transactions_cols, "fichier des transactions")
        validate_required_columns(df_journal, required_journal_cols, "fichier journal")
        
        print("3. Nettoyage et formatage des données...")
        
        # Nettoyage des colonnes de texte utilisées comme clés de jointure
        df_orders = clean_text_data(df_orders, ['Name', 'Billing name', 'Financial Status', 'Payment Method'])
        df_transactions = clean_text_data(df_transactions, ['Order', 'Payment Method Name'])
        df_journal = clean_text_data(df_journal, ['Piece', 'Référence LMB'])
        
        # Formatage des dates - conversion en format français jj/mm/aaaa
        df_orders['Fulfilled at'] = df_orders['Fulfilled at'].apply(format_date_to_french)
          # Formatage des colonnes monétaires en type numérique
        monetary_cols_orders = ['Tax 1 Value', 'Outstanding Balance']
        monetary_cols_transactions = ['Presentment Amount', 'Fee', 'Net']
        
        for col in monetary_cols_orders:
            if col in df_orders.columns:
                df_orders[col] = pd.to_numeric(df_orders[col], errors='coerce').fillna(0)
        
        for col in monetary_cols_transactions:
            if col in df_transactions.columns:
                df_transactions[col] = pd.to_numeric(df_transactions[col], errors='coerce').fillna(0)
        
        print("3.5. Agrégation des commandes pour éviter les doublons...")
        
        # IMPORTANT: Grouper les commandes par Name pour éviter les doublons
        # (cas où il y a plusieurs lignes de produits par commande)
        print(f"   - Nombre de lignes avant agrégation des commandes: {len(df_orders)}")
        
        # Colonnes à garder telles quelles (prendre la première valeur)
        first_value_cols = ['Fulfilled at', 'Billing name', 'Financial Status', 'Payment Method']
        
        # Colonnes à sommer (montants, quantités, etc.)
        sum_cols = ['Tax 1 Value', 'Outstanding Balance']
          # Détecter automatiquement d'autres colonnes numériques qui pourraient nécessiter une sommation        # Temporairement désactiver l'auto-détection pour éviter les erreurs pandas
        # for col in df_orders.columns:
        #     if col not in first_value_cols and col != 'Name':
        #         try:
        #             # Si c'est une colonne numérique, on la somme
        #             if pd.api.types.is_numeric_dtype(df_orders[col]):
        #                 if col not in sum_cols:
        #                     sum_cols.append(col)
        #                     print(f"   - Colonne numérique détectée pour sommation: {col}")
        #             # Sinon, on prend la première valeur
        #             elif col not in first_value_cols:
        #                 first_value_cols.append(col)
        #         except Exception as e:
        #             print(f"   - Erreur lors de l'analyse de la colonne {col}: {e}")
        #             # En cas d'erreur, traiter comme une colonne de première valeur
        #             if col not in first_value_cols:
        #                 first_value_cols.append(col)
        
        # Ajouter toutes les autres colonnes aux first_value_cols pour éviter les erreurs
        for col in df_orders.columns:
            if col not in first_value_cols and col not in sum_cols and col != 'Name':
                first_value_cols.append(col)
        
        # Préparer les opérations d'agrégation
        agg_operations = {}
        
        # Pour les colonnes de texte/dates, prendre la première valeur
        for col in first_value_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'first'
          # Pour les colonnes monétaires, faire la somme
        for col in sum_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'sum'
        
        # Grouper par Name (identifiant unique de la commande)
        if agg_operations:
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        else:
            # Si pas d'opérations d'agrégation, juste dédupliquer
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        
        print(f"   - Nombre de lignes après agrégation des commandes: {len(df_orders_aggregated)}")
        
        # Note: Un même client peut avoir plusieurs commandes distinctes
        # Chaque commande doit apparaître sur une ligne séparée
        
        # Remplacer df_orders par la version agrégée
        df_orders = df_orders_aggregated
          # ÉTAPE 2: Agrégation des transactions par commande
        print("4. Agrégation des transactions par commande...")
        
        # Grouper par Order et sommer les montants pour éviter les doublons
        # IMPORTANT: Garder aussi Payment Method Name (prendre la première valeur)
        df_transactions_aggregated = df_transactions.groupby('Order').agg({
            'Presentment Amount': 'sum',
            'Fee': 'sum',
            'Net': 'sum',
            'Payment Method Name': 'first'  # Garder la méthode de paiement
        }).reset_index()
        
        print(f"   - Transactions après agrégation: {len(df_transactions_aggregated)} lignes")
          # ÉTAPE 3: Fusion des DataFrames
        print("5. Fusion des DataFrames...")
        
        # Première fusion: Commandes + Transactions agrégées (jointure à gauche)
        df_merged_step1 = pd.merge(df_orders, df_transactions_aggregated, 
                                  left_on='Name', right_on='Order', how='left')
        print(f"   - Après fusion commandes-transactions: {len(df_merged_step1)} lignes")
        
        # Diagnostic avant fusion avec journal
        print("   - Diagnostic avant fusion avec journal:")
        print(f"     * Commandes uniques dans df_merged_step1: {df_merged_step1['Name'].nunique()} ({list(df_merged_step1['Name'].unique()[:5])}...)")
        print(f"     * Références uniques dans journal: {df_journal['Piece'].nunique()} ({list(df_journal['Piece'].unique()[:5])}...)")
          # Vérifier les correspondances
        commandes_dans_journal = df_merged_step1['Name'].isin(df_journal['Piece']).sum()
        print(f"     * Commandes qui ont une correspondance dans le journal: {commandes_dans_journal}/{len(df_merged_step1)}")
          # Deuxième fusion: Résultat + Journal (jointure à gauche)
        # TOUJOURS essayer d'améliorer les correspondances avec normalisation
        print("DEBUG: Début de la logique de fusion avec journal")
        print(f"DEBUG: commandes_dans_journal = {commandes_dans_journal}, len(df_merged_step1) = {len(df_merged_step1)}")
        if commandes_dans_journal < len(df_merged_step1):  # Si pas 100% de correspondances
            print("     🔧 Application de la normalisation des références...")
            df_merged_step1_improved = improve_journal_matching(df_merged_step1, df_journal)
            
            # Utiliser les données améliorées
            df_merged_final = df_merged_step1_improved
        else:
            print("     ✅ Toutes les correspondances trouvées, fusion standard")
            # Fusion standard si toutes les correspondances sont déjà trouvées
            df_merged_final = pd.merge(df_merged_step1, df_journal, 
                                      left_on='Name', right_on='Piece', how='left')
        print(f"   - Après fusion avec journal: {len(df_merged_final)} lignes")
        
        # Diagnostic après fusion
        ref_lmb_non_nulles = df_merged_final['Référence LMB'].notna().sum()
        print(f"   - Références LMB trouvées: {ref_lmb_non_nulles}/{len(df_merged_final)} ({ref_lmb_non_nulles/len(df_merged_final)*100:.1f}%)")        # ÉTAPE 4: Création du tableau final avec les 16 colonnes
        print("6. Création du tableau final...")
        
        # DEBUG: Vérifier les colonnes disponibles dans df_merged_final
        print(f"DEBUG: Colonnes dans df_merged_final: {list(df_merged_final.columns)}")
        print(f"DEBUG: 'Payment Method Name' présente: {'Payment Method Name' in df_merged_final.columns}")
        
        # DEBUG: Vérifier quelques exemples de Payment Method Name
        if 'Payment Method Name' in df_merged_final.columns:
            payment_methods = df_merged_final['Payment Method Name'].dropna().unique()
            print(f"DEBUG: Valeurs uniques de Payment Method Name: {payment_methods}")
        
        df_final = pd.DataFrame()
        
        # Colonnes du tableau final dans l'ordre requis
        df_final['Centre de profit'] = 'lcdi.fr'  # Valeur statique
        df_final['Réf.WEB'] = df_merged_final['Name']
        df_final['Réf. LMB'] = df_merged_final['Référence LMB'].fillna('')
        df_final['Date Facture'] = calculate_invoice_dates(df_merged_final)
        df_final['Etat'] = df_merged_final['Financial Status'].fillna('').apply(translate_financial_status)
        df_final['Client'] = df_merged_final['Billing name'].fillna('')
        
        # Calculs des montants
        corrected_amounts = calculate_corrected_amounts(df_merged_final)
        df_final['HT'] = corrected_amounts['HT']
        df_final['TVA'] = corrected_amounts['TVA']
        df_final['TTC'] = corrected_amounts['TTC']
        df_final['reste'] = df_merged_final['Outstanding Balance'].fillna(0)
        df_final['Shopify'] = df_merged_final['Net'].fillna(0)
        df_final['Frais de commission'] = df_merged_final['Fee'].fillna(0)
        
        # Traitement des méthodes de paiement
        print("7. Traitement des méthodes de paiement...")
        payment_categorization = df_merged_final.apply(
            lambda row: categorize_payment_method(
                row.get('Payment Method'),  # Méthode de paiement des commandes
                row.get('Payment Method Name'),  # Méthode de paiement des transactions (plus précise pour PayPal),
                corrected_amounts['TTC'].loc[row.name] if row.name in corrected_amounts['TTC'].index else None,  # Utiliser le TTC calculé
                fallback_amount=row.get('Total', 0)  # Fallback sur le montant de la commande
            ), 
            axis=1
        )
        
        df_final['Virement bancaire'] = [pm['Virement bancaire'] for pm in payment_categorization]
        df_final['Carte bancaire'] = [pm['Carte bancaire'] for pm in payment_categorization]
        df_final['ALMA'] = [pm['ALMA'] for pm in payment_categorization]
        df_final['Younited'] = [pm['Younited'] for pm in payment_categorization]
        df_final['PayPal'] = [pm['PayPal'] for pm in payment_categorization]          # PRÉPARATION STATUT DYNAMIQUE: Créer une colonne vide pour les formules Excel
        # Les formules seront ajoutées lors de la génération du fichier Excel
        df_final['Statut'] = ''  # Colonne vide pour les formules
        
        print("8. Nettoyage final des données...")
        
        # Appliquer les indicateurs d'informations manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        # S'assurer que "Centre de profit" est toujours "lcdi.fr" (forcer après toutes les fusions)
        df_final['Centre de profit'] = 'lcdi.fr'
        
        # Indicateurs de données manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        print(f"=== TRAITEMENT TERMINÉ ===")
        print(f"Tableau final généré avec {len(df_final)} lignes et {len(df_final.columns)} colonnes")
        
        return df_final
        
    except Exception as e:
        print(f"ERREUR lors du traitement: {str(e)}")
        raise e

def process_dataframes_directly(df_orders, df_transactions, df_journal):
    """
    Fonction auxiliaire pour traiter directement des DataFrames (utilisée pour les tests)
    """
    try:
        print("=== DÉBUT DU TRAITEMENT (DataFrames) ===")
        
        # ÉTAPE 1: Les DataFrames sont déjà chargés
        print("1. DataFrames déjà chargés...")
        print(f"   - Commandes: {len(df_orders)} lignes")
        print(f"   - Transactions: {len(df_transactions)} lignes")
        print(f"   - Journal: {len(df_journal)} lignes")
          # ÉTAPE 2: Vérification et normalisation des colonnes
        print("2. Vérification et normalisation des colonnes...")
          # Normaliser les noms de colonnes pour les commandes
        required_orders_cols = ['Name', 'Fulfilled at', 'Billing name', 'Financial Status', 'Tax 1 Value', 'Outstanding Balance', 'Payment Method', 'Total', 'Taxes']
        df_orders = normalize_column_names(df_orders, required_orders_cols, 'commandes')
        validate_required_columns(df_orders, required_orders_cols, "fichier des commandes")
        
        # Normaliser les noms de colonnes pour les transactions
        required_trans_cols = ['Order', 'Presentment Amount', 'Fee', 'Net']
        df_transactions = normalize_column_names(df_transactions, required_trans_cols, 'transactions')
        validate_required_columns(df_transactions, required_trans_cols, "fichier des transactions")
        
        # Normaliser les noms de colonnes pour le journal
        required_journal_cols = ['Piece', 'Référence LMB']
        df_journal = normalize_column_names(df_journal, required_journal_cols, 'journal')
        validate_required_columns(df_journal, required_journal_cols, "fichier journal")
          # ÉTAPE 3: Nettoyage et formatage des données
        print("3. Nettoyage et formatage des données...")
        
        # Nettoyage des colonnes de texte utilisées comme clés de jointure
        df_orders = clean_text_data(df_orders, ['Name', 'Billing name', 'Financial Status', 'Payment Method'])
        df_transactions = clean_text_data(df_transactions, ['Order', 'Payment Method Name'])
        df_journal = clean_text_data(df_journal, ['Piece', 'Référence LMB'])
        
        # Formatage des dates - conversion en format français jj/mm/aaaa
        df_orders['Fulfilled at'] = df_orders['Fulfilled at'].apply(format_date_to_french)
        
        # Formatage des colonnes monétaires en type numérique
        monetary_cols_orders = ['Tax 1 Value', 'Outstanding Balance']
        monetary_cols_transactions = ['Presentment Amount', 'Fee', 'Net']
        
        for col in monetary_cols_orders:
            if col in df_orders.columns:
                df_orders[col] = pd.to_numeric(df_orders[col], errors='coerce').fillna(0)
        
        for col in monetary_cols_transactions:
            if col in df_transactions.columns:
                df_transactions[col] = pd.to_numeric(df_transactions[col], errors='coerce').fillna(0)
        
        # ÉTAPE 3.5: Agrégation des commandes pour éviter les doublons par commande
        print("3.5. Agrégation des commandes pour éviter les doublons...")
        print(f"   - Nombre de lignes avant agrégation des commandes: {len(df_orders)}")
        
        # Définir les colonnes pour l'agrégation        # Listes de base pour l'agrégation
        first_cols = ['Fulfilled at', 'Billing name', 'Financial Status', 'Payment Method', 'Email', 'Lineitem name']
        sum_cols = ['Tax 1 Value', 'Outstanding Balance']
        
        # Colonnes monétaires spécifiques à sommer (éviter l'auto-détection problématique)
        predefined_sum_cols = ['Total', 'Taxes', 'Shipping', 'Discount Amount', 'Refunded Amount', 
                              'Lineitem price', 'Lineitem quantity', 'Outstanding Balance']
        
        for col in predefined_sum_cols:
            if col in df_orders.columns and col not in sum_cols:
                try:
                    # Vérifier que la colonne contient réellement des valeurs numériques
                    non_null_values = df_orders[col].dropna()
                    if len(non_null_values) > 0:
                        # Tenter de convertir en numérique
                        pd.to_numeric(non_null_values, errors='raise')
                        sum_cols.append(col)
                        print(f"   - Colonne prédéfinie ajoutée pour sommation: {col}")
                except Exception:
                    # Si conversion échoue, traiter comme première valeur
                    if col not in first_cols:
                        first_cols.append(col)
                        print(f"   - Colonne {col} traitée comme 'first' (non numérique)")
        
        # Ajouter les autres colonnes non numériques à first_cols
        for col in df_orders.columns:
            if col not in sum_cols and col not in first_cols and col != 'Name':
                first_cols.append(col)
        
        # Construire le dictionnaire d'opérations d'agrégation
        agg_operations = {}
        
        # Configurer les opérations d'agrégation
        for col in first_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'first'
        
        # Pour les colonnes monétaires, faire la somme
        for col in sum_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'sum'
        
        # Grouper par Name (identifiant unique de la commande)
        if agg_operations:
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        else:
            # Si pas d'opérations d'agrégation, juste dédupliquer
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        
        print(f"   - Nombre de lignes après agrégation des commandes: {len(df_orders_aggregated)}")
        
        # Note: Un même client peut avoir plusieurs commandes distinctes
        # Chaque commande doit apparaître sur une ligne séparée
        
        # Remplacer df_orders par la version agrégée
        df_orders = df_orders_aggregated
        
        # ÉTAPE 4: Agrégation des transactions par commande
        print("4. Agrégation des transactions par commande...")
        
        # Grouper par Order et sommer les montants pour éviter les doublons
        df_transactions_aggregated = df_transactions.groupby('Order').agg({
            'Presentment Amount': 'sum',
            'Fee': 'sum',
            'Net': 'sum',
            'Payment Method Name': 'first'  # Garder la méthode de paiement
        }).reset_index()
        
        print(f"   - Transactions après agrégation: {len(df_transactions_aggregated)} lignes")
          # ÉTAPE 5: Fusion des DataFrames
        print("5. Fusion des DataFrames...")
        
        # Première fusion: Commandes + Transactions agrégées (jointure à gauche)
        df_merged_step1 = pd.merge(df_orders, df_transactions_aggregated, 
                                  left_on='Name', right_on='Order', how='left')
        print(f"   - Après fusion commandes-transactions: {len(df_merged_step1)} lignes")
        
        # Diagnostic avant fusion avec journal
        print("   - Diagnostic avant fusion avec journal:")
        print(f"     * Commandes uniques dans df_merged_step1: {df_merged_step1['Name'].nunique()} ({list(df_merged_step1['Name'].unique()[:5])}...)")
        print(f"     * Références uniques dans journal: {df_journal['Piece'].nunique()} ({list(df_journal['Piece'].unique()[:5])}...)")
          # Vérifier les correspondances
        commandes_dans_journal = df_merged_step1['Name'].isin(df_journal['Piece']).sum()
        print(f"     * Commandes qui ont une correspondance dans le journal: {commandes_dans_journal}/{len(df_merged_step1)}")
        
        # Deuxième fusion: Résultat + Journal (jointure à gauche)
        # TOUJOURS essayer d'améliorer les correspondances avec normalisation
        print("DEBUG: Version 2 - Début de la logique de fusion avec journal")
        print(f"DEBUG: Version 2 - commandes_dans_journal = {commandes_dans_journal}, len(df_merged_step1) = {len(df_merged_step1)}")
        if commandes_dans_journal < len(df_merged_step1):  # Si pas 100% de correspondances
            print("     🔧 [V2] Application de la normalisation des références...")
            df_merged_step1_improved = improve_journal_matching(df_merged_step1, df_journal)
            
            # Utiliser les données améliorées
            df_merged_final = df_merged_step1_improved
        else:
            print("     ✅ [V2] Toutes les correspondances trouvées, fusion standard")
            # Fusion standard si toutes les correspondances sont déjà trouvées
            df_merged_final = pd.merge(df_merged_step1, df_journal, 
                                      left_on='Name', right_on='Piece', how='left')
        print(f"   - Après fusion avec journal: {len(df_merged_final)} lignes")
        
        # Diagnostic après fusion
        ref_lmb_non_nulles = df_merged_final['Référence LMB'].notna().sum()
        print(f"   - Références LMB trouvées: {ref_lmb_non_nulles}/{len(df_merged_final)} ({ref_lmb_non_nulles/len(df_merged_final)*100:.1f}%)")
          # ÉTAPE 6: Création du tableau final avec les 16 colonnes
        print("6. Création du tableau final...")
        
        df_final = pd.DataFrame()
        
        # Colonnes du tableau final dans l'ordre requis
        df_final['Centre de profit'] = 'lcdi.fr'  # Valeur statique
        df_final['Réf.WEB'] = df_merged_final['Name']
        df_final['Réf. LMB'] = df_merged_final['Référence LMB'].fillna('')
        df_final['Date Facture'] = calculate_invoice_dates(df_merged_final)
        df_final['Etat'] = df_merged_final['Financial Status'].fillna('').apply(translate_financial_status)
        df_final['Client'] = df_merged_final['Billing name'].fillna('')
        
        # Calculs des montants
        corrected_amounts = calculate_corrected_amounts(df_merged_final)
        df_final['HT'] = corrected_amounts['HT']
        df_final['TVA'] = corrected_amounts['TVA']
        df_final['TTC'] = corrected_amounts['TTC']
        df_final['reste'] = df_merged_final['Outstanding Balance'].fillna(0)
        df_final['Shopify'] = df_merged_final['Net'].fillna(0)
        df_final['Frais de commission'] = df_merged_final['Fee'].fillna(0)
        
        # Traitement des méthodes de paiement
        print("7. Traitement des méthodes de paiement...")
        payment_categorization = df_merged_final.apply(
            lambda row: categorize_payment_method(
                row.get('Payment Method'),  # Méthode de paiement des commandes
                row.get('Payment Method Name'),  # Méthode de paiement des transactions (plus précise pour PayPal),
                corrected_amounts['TTC'].loc[row.name] if row.name in corrected_amounts['TTC'].index else None,  # Utiliser le TTC calculé
                fallback_amount=row.get('Total', 0)  # Fallback sur le montant de la commande
            ), 
            axis=1
        )
        
        df_final['Virement bancaire'] = [pm['Virement bancaire'] for pm in payment_categorization]
        df_final['Carte bancaire'] = [pm['Carte bancaire'] for pm in payment_categorization]
        df_final['ALMA'] = [pm['ALMA'] for pm in payment_categorization]
        df_final['Younited'] = [pm['Younited'] for pm in payment_categorization]
        df_final['PayPal'] = [pm['PayPal'] for pm in payment_categorization]          # PRÉPARATION STATUT DYNAMIQUE: Créer une colonne vide pour les formules Excel
        # Les formules seront ajoutées lors de la génération du fichier Excel
        df_final['Statut'] = ''  # Colonne vide pour les formules
        
        print("8. Nettoyage final des données...")
        
        # Appliquer les indicateurs d'informations manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        # S'assurer que "Centre de profit" est toujours "lcdi.fr" (forcer après toutes les fusions)
        df_final['Centre de profit'] = 'lcdi.fr'
        
        # Indicateurs de données manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        print(f"=== TRAITEMENT TERMINÉ ===")
        print(f"Tableau final généré avec {len(df_final)} lignes et {len(df_final.columns)} colonnes")
        
        return df_final
        
    except Exception as e:
        print(f"ERREUR lors du traitement: {str(e)}")
        raise e

def process_dataframes_with_normalization(df_orders, df_transactions, df_journal):
    """
    Version améliorée qui utilise toujours la normalisation des références
    """
    try:
        print("=== DÉBUT DU TRAITEMENT AVEC NORMALISATION ===")
        
        # ÉTAPE 1: Les DataFrames sont déjà chargés
        print("1. DataFrames déjà chargés...")
        print(f"   - Commandes: {len(df_orders)} lignes")
        print(f"   - Transactions: {len(df_transactions)} lignes")
        print(f"   - Journal: {len(df_journal)} lignes")
          # ÉTAPE 2: Vérification et normalisation des colonnes
        print("2. Vérification et normalisation des colonnes...")
          # Définir les colonnes requises
        required_orders_cols = ['Name', 'Fulfilled at', 'Billing name', 'Financial Status', 'Tax 1 Value', 'Outstanding Balance', 'Payment Method', 'Total', 'Taxes']
        required_transactions_cols = ['Order', 'Presentment Amount', 'Fee', 'Net', 'Payment Method Name']
        required_journal_cols = ['Piece', 'Référence LMB']
          # Normaliser les noms de colonnes pour les commandes
        df_orders = normalize_column_names(df_orders, required_orders_cols, 'commandes')
        validate_required_columns(df_orders, required_orders_cols, "fichier des commandes")
        
        # Normaliser les noms de colonnes pour les transactions
        df_transactions = normalize_column_names(df_transactions, required_transactions_cols, 'transactions')
        validate_required_columns(df_transactions, required_transactions_cols, "fichier des transactions")
        
        # Normaliser les noms de colonnes pour le journal
        df_journal = normalize_column_names(df_journal, required_journal_cols, 'journal')
        validate_required_columns(df_journal, required_journal_cols, "fichier journal")
        
        # ÉTAPE 3: Nettoyage et formatage des données
        print("3. Nettoyage et formatage des données...")
        
        # Nettoyage des colonnes de texte utilisées comme clés de jointure
        df_orders = clean_text_data(df_orders, ['Name', 'Billing name', 'Financial Status', 'Payment Method'])
        df_transactions = clean_text_data(df_transactions, ['Order', 'Payment Method Name'])
        df_journal = clean_text_data(df_journal, ['Piece', 'Référence LMB'])
        
        # Formatage des dates - conversion en format français jj/mm/aaaa
        df_orders['Fulfilled at'] = df_orders['Fulfilled at'].apply(format_date_to_french)
        
        # Formatage des colonnes monétaires en type numérique
        monetary_cols_orders = ['Tax 1 Value', 'Outstanding Balance']
        monetary_cols_transactions = ['Presentment Amount', 'Fee', 'Net']
        
        for col in monetary_cols_orders:
            if col in df_orders.columns:
                df_orders[col] = pd.to_numeric(df_orders[col], errors='coerce').fillna(0)
        
        for col in monetary_cols_transactions:
            if col in df_transactions.columns:
                df_transactions[col] = pd.to_numeric(df_transactions[col], errors='coerce').fillna(0)
        
        # ÉTAPE 3.5: Agrégation des commandes pour éviter les doublons
        print("3.5. Agrégation des commandes pour éviter les doublons...")
        print(f"   - Nombre de lignes avant agrégation des commandes: {len(df_orders)}")
        
        # Définir les colonnes pour l'agrégation
        first_cols = ['Fulfilled at', 'Billing name', 'Financial Status', 'Payment Method', 'Email', 'Lineitem name']
        sum_cols = ['Tax 1 Value', 'Outstanding Balance']
        
        # Construire le dictionnaire d'opérations d'agrégation
        agg_operations = {}
        
        # Détecter automatiquement les colonnes numériques pour la sommation
        for col in df_orders.columns:
            if col not in first_cols and col != 'Name':  # Name est la clé de groupement
                if df_orders[col].dtype in ['int64', 'float64'] or pd.api.types.is_numeric_dtype(df_orders[col]):
                    if col not in sum_cols:
                        sum_cols.append(col)
                        print(f"   - Colonne numérique détectée pour sommation: {col}")
                elif col not in first_cols:
                    first_cols.append(col)
        
        # Configurer les opérations d'agrégation
        for col in first_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'first'
        
        # Pour les colonnes monétaires, faire la somme
        for col in sum_cols:
            if col in df_orders.columns:
                agg_operations[col] = 'sum'
        
        # Grouper par Name (identifiant unique de la commande)
        if agg_operations:
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        else:
            # Si pas d'opérations d'agrégation, juste dédupliquer
            df_orders_aggregated = df_orders.drop_duplicates(subset=['Name'], keep='first')
        
        print(f"   - Nombre de lignes après agrégation des commandes: {len(df_orders_aggregated)}")
        
        # Note: Un même client peut avoir plusieurs commandes distinctes
        # Chaque commande doit apparaître sur une ligne séparée
        
        # Remplacer df_orders par la version agrégée
        df_orders = df_orders_aggregated
        
        # ÉTAPE 4: Agrégation des transactions par commande
        print("4. Agrégation des transactions par commande...")
        
        # Grouper par Order et sommer les montants pour éviter les doublons
        df_transactions_aggregated = df_transactions.groupby('Order').agg({
            'Presentment Amount': 'sum',
            'Fee': 'sum',
            'Net': 'sum',
            'Payment Method Name': 'first'  # Garder la méthode de paiement
        }).reset_index()
        
        print(f"   - Transactions après agrégation: {len(df_transactions_aggregated)} lignes")
          # ÉTAPE 5: Fusion des DataFrames
        print("5. Fusion des DataFrames...")
        
        # Première fusion: Commandes + Transactions agrégées (jointure à gauche)
        df_merged_step1 = pd.merge(df_orders, df_transactions_aggregated, 
                                  left_on='Name', right_on='Order', how='left')
        print(f"   - Après fusion commandes-transactions: {len(df_merged_step1)} lignes")
        
        # Diagnostic avant fusion avec journal
        print("   - Diagnostic avant fusion avec journal:")
        print(f"     * Commandes uniques dans df_merged_step1: {df_merged_step1['Name'].nunique()} ({list(df_merged_step1['Name'].unique()[:5])}...)")
        print(f"     * Références uniques dans journal: {df_journal['Piece'].nunique()} ({list(df_journal['Piece'].unique()[:5])}...)")
          # Vérifier les correspondances
        commandes_dans_journal = df_merged_step1['Name'].isin(df_journal['Piece']).sum()
        print(f"     * Commandes qui ont une correspondance dans le journal: {commandes_dans_journal}/{len(df_merged_step1)}")
        
        # Deuxième fusion: Résultat + Journal (jointure à gauche)
        # TOUJOURS essayer d'améliorer les correspondances avec normalisation
        print("DEBUG: Version 2 - Début de la logique de fusion avec journal")
        print(f"DEBUG: Version 2 - commandes_dans_journal = {commandes_dans_journal}, len(df_merged_step1) = {len(df_merged_step1)}")
        if commandes_dans_journal < len(df_merged_step1):  # Si pas 100% de correspondances
            print("     🔧 [V2] Application de la normalisation des références...")
            df_merged_step1_improved = improve_journal_matching(df_merged_step1, df_journal)
            
            # Utiliser les données améliorées
            df_merged_final = df_merged_step1_improved
        else:
            print("     ✅ [V2] Toutes les correspondances trouvées, fusion standard")
            # Fusion standard si toutes les correspondances sont déjà trouvées
            df_merged_final = pd.merge(df_merged_step1, df_journal, 
                                      left_on='Name', right_on='Piece', how='left')
        print(f"   - Après fusion avec journal: {len(df_merged_final)} lignes")
        
        # Diagnostic après fusion
        ref_lmb_non_nulles = df_merged_final['Référence LMB'].notna().sum()
        print(f"   - Références LMB trouvées: {ref_lmb_non_nulles}/{len(df_merged_final)} ({ref_lmb_non_nulles/len(df_merged_final)*100:.1f}%)")
          # ÉTAPE 6: Création du tableau final avec les 16 colonnes
        print("6. Création du tableau final...")
        
        df_final = pd.DataFrame()
        
        # Colonnes du tableau final dans l'ordre requis
        df_final['Centre de profit'] = 'lcdi.fr'  # Valeur statique
        df_final['Réf.WEB'] = df_merged_final['Name']
        df_final['Réf. LMB'] = df_merged_final['Référence LMB'].fillna('')
        df_final['Date Facture'] = calculate_invoice_dates(df_merged_final)
        df_final['Etat'] = df_merged_final['Financial Status'].fillna('').apply(translate_financial_status)
        df_final['Client'] = df_merged_final['Billing name'].fillna('')
          # Calculs des montants
        corrected_amounts = calculate_corrected_amounts(df_merged_final)
        df_final['HT'] = corrected_amounts['HT']
        df_final['TVA'] = corrected_amounts['TVA']
        df_final['TTC'] = corrected_amounts['TTC']
        df_final['reste'] = df_merged_final['Outstanding Balance'].fillna(0)
        df_final['Shopify'] = df_merged_final['Net'].fillna(0)
        df_final['Frais de commission'] = df_merged_final['Fee'].fillna(0)
        
        # Traitement des méthodes de paiement
        print("7. Traitement des méthodes de paiement...")
        payment_categorization = df_merged_final.apply(
            lambda row: categorize_payment_method(
                row.get('Payment Method'),  # Méthode de paiement des commandes
                row.get('Payment Method Name'),  # Méthode de paiement des transactions (plus précise pour PayPal),
                corrected_amounts['TTC'].loc[row.name] if row.name in corrected_amounts['TTC'].index else None,  # Utiliser le TTC calculé
                fallback_amount=row.get('Total', 0)  # Fallback sur le montant de la commande
            ), 
            axis=1
        )
        
        df_final['Virement bancaire'] = [pm['Virement bancaire'] for pm in payment_categorization]
        df_final['Carte bancaire'] = [pm['Carte bancaire'] for pm in payment_categorization]
        df_final['ALMA'] = [pm['ALMA'] for pm in payment_categorization]
        df_final['Younited'] = [pm['Younited'] for pm in payment_categorization]
        df_final['PayPal'] = [pm['PayPal'] for pm in payment_categorization]          # PRÉPARATION STATUT DYNAMIQUE: Créer une colonne vide pour les formules Excel
        # Les formules seront ajoutées lors de la génération du fichier Excel
        df_final['Statut'] = ''  # Colonne vide pour les formules
        
        print("8. Nettoyage final des données...")
        
        # Appliquer les indicateurs d'informations manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        # S'assurer que "Centre de profit" est toujours "lcdi.fr" (forcer après toutes les fusions)
        df_final['Centre de profit'] = 'lcdi.fr'
        
        # Indicateurs de données manquantes
        df_final = fill_missing_data_indicators(df_final, df_merged_final)
        
        print(f"=== TRAITEMENT TERMINÉ ===")
        print(f"Tableau final généré avec {len(df_final)} lignes et {len(df_final.columns)} colonnes")
        
        return df_final
        
    except Exception as e:
        print(f"ERREUR lors du traitement: {str(e)}")
        raise e

def translate_financial_status(status):
    """
    Traduit les statuts financiers anglais en français
    """
    if pd.isna(status) or status == '':
        return ''
    
    status_str = str(status).lower().strip()
    
    # Dictionnaire de traduction
    translations = {
        'paid': 'payée',
        'pending': 'en attente',
        'partially_paid': 'payée partiellement',
        'refunded': 'remboursée',
        'partially_refunded': 'remboursée partiellement',
        'voided': 'annulée'
    }
    
    return translations.get(status_str, status_str)

def normalize_reference_format(ref):
    """
    Normalise le format des références pour améliorer les correspondances
    Gère les formats LCDI-XXXX, #LCDI-XXXX, #lcdi-xxxx, etc.
    """
    if pd.isna(ref) or ref == '':
        return ref
    
    ref_str = str(ref).strip().upper()
    
    # Pattern pour capturer LCDI-XXXX avec ou sans #
    lcdi_pattern = r'#?LCDI[-_]?(\d+)'
    matches = re.findall(lcdi_pattern, ref_str)
    
    if matches:
        # Retourner au format standard #LCDI-XXXX
        return f"#LCDI-{matches[0]}"
    
    # Si pas de pattern LCDI, essayer de capturer juste les chiffres
    numbers = re.findall(r'\d+', ref_str)
    if numbers:
        return f"#{numbers[0]}"
    
    return ref_str

def normalize_reference_with_multiples(ref):
    """
    Normalise les références en gérant les cas de références multiples
    Ex: "LCDI-1020 LCDI-1021" -> ["#LCDI-1020", "#LCDI-1021"]
    """
    if pd.isna(ref) or ref == '':
        return []
    
    ref_str = str(ref).strip().upper()
    
    # Pattern pour capturer toutes les références LCDI-XXXX
    lcdi_pattern = r'#?LCDI[-_]?(\d+)'
    matches = re.findall(lcdi_pattern, ref_str)
    
    if matches:
        return [f"#LCDI-{match}" for match in matches]
    
    # Si pas de pattern LCDI, essayer de capturer juste les chiffres
    numbers = re.findall(r'\d+', ref_str)
    if numbers:
        return [f"#{num}" for num in numbers]
    
    return [ref_str]

def improve_journal_matching(df_orders, df_journal):
    """
    Fusion améliorée avec gestion des références multiples
    Gère les formats #LCDI-XXXX vs LCDI-XXXX et les références multiples comme 'LCDI-1020 LCDI-1021'
    """
    print("   - Fusion avec normalisation et gestion des références multiples...")
    
    # Copier les DataFrames pour éviter de modifier les originaux
    df_orders_copy = df_orders.copy()
    df_journal_copy = df_journal.copy()
    
    # Trouver la colonne de référence dans le journal (peut être 'Piece' après normalisation)
    journal_ref_col = 'Piece'  # Nom standardisé après normalize_column_names
    
    if journal_ref_col not in df_journal_copy.columns:
        print(f"❌ Erreur: Colonne '{journal_ref_col}' non trouvée dans le journal")
        print(f"Colonnes disponibles: {list(df_journal_copy.columns)}")
        return df_orders_copy  # Retourner les commandes sans fusion
    
    # Normaliser les références des commandes : toujours au format #LCDI-XXXX
    df_orders_copy['Name_normalized'] = df_orders_copy['Name'].apply(
        lambda x: x if str(x).startswith('#') else f"#{x}" if pd.notna(x) else None
    )
    
    # Créer un dictionnaire de mapping : référence normalisée -> données journal
    journal_mapping = {}
    
    # Pour chaque ligne du journal
    for journal_idx, journal_row in df_journal_copy.iterrows():
        journal_ref = journal_row[journal_ref_col]
        if pd.isna(journal_ref):
            continue
            
        journal_ref_str = str(journal_ref).strip()
        
        # Cas 1: Référence simple (ex: LCDI-1038 ou #LCDI-1038)
        if ' ' not in journal_ref_str:
            # Normaliser la référence
            journal_normalized = journal_ref_str if journal_ref_str.startswith('#') else f"#{journal_ref_str}"
            journal_mapping[journal_normalized] = journal_row          # Cas 2: Référence multiple (ex: LCDI-1020 LCDI-1021)
        else:
            import re
            # Extraire tous les numéros de commandes
            numbers = re.findall(r'LCDI-(\d+)', journal_ref_str)
            
            if numbers:
                print(f"     - Référence multiple détectée: '{journal_ref_str}' -> commandes {numbers}")
                
                # Pour les références multiples, on doit répartir les montants
                # Stratégie: calculer le poids de chaque commande et répartir proportionnellement
                
                # Récupérer les montants totaux des commandes concernées pour calculer les proportions
                command_totals = {}
                total_sum = 0
                
                for num in numbers:
                    target_ref = f"#LCDI-{num}"
                    # Trouver la commande correspondante dans df_orders_copy
                    matching_orders = df_orders_copy[df_orders_copy['Name_normalized'] == target_ref]
                    if not matching_orders.empty:
                        order_total = pd.to_numeric(matching_orders.iloc[0]['Total'], errors='coerce')
                        if pd.notna(order_total):
                            command_totals[target_ref] = order_total
                            total_sum += order_total
                        else:
                            command_totals[target_ref] = 0
                    else:
                        command_totals[target_ref] = 0
                
                print(f"       - Totaux des commandes : {command_totals}, somme: {total_sum}")
                
                # Récupérer les montants du journal
                journal_ttc = journal_row.get('Montant du document TTC', None)
                journal_ht = journal_row.get('Montant du document HT', None)
                journal_marge = journal_row.get('Montant marge HT', None)
                
                # Convertir les montants du journal au format numérique
                if pd.notna(journal_ttc):
                    try:
                        journal_ttc_num = float(str(journal_ttc).replace(',', '.').replace(' ', ''))
                    except:
                        journal_ttc_num = None
                else:
                    journal_ttc_num = None
                    
                if pd.notna(journal_ht):
                    try:
                        journal_ht_num = float(str(journal_ht).replace(',', '.').replace(' ', ''))
                    except:
                        journal_ht_num = None
                else:
                    journal_ht_num = None
                    
                if pd.notna(journal_marge):
                    try:
                        journal_marge_num = float(str(journal_marge).replace(',', '.').replace(' ', ''))
                    except:
                        journal_marge_num = None
                else:
                    journal_marge_num = None
                
                print(f"       - Montants journal : TTC={journal_ttc_num}, HT={journal_ht_num}, Marge={journal_marge_num}")
                
                # Répartir les montants proportionnellement
                for num in numbers:
                    target_ref = f"#LCDI-{num}"
                    
                    # Créer une copie de la ligne journal pour cette commande
                    proportional_journal_data = journal_row.copy()
                    
                    # Calculer la proportion de cette commande
                    if total_sum > 0 and command_totals[target_ref] > 0:
                        proportion = command_totals[target_ref] / total_sum
                        print(f"       - {target_ref}: proportion = {proportion:.3f}")
                        
                        # Répartir les montants
                        if journal_ttc_num is not None:
                            proportional_ttc = journal_ttc_num * proportion
                            proportional_journal_data['Montant du document TTC'] = f"{proportional_ttc:.2f}".replace('.', ',')
                        
                        if journal_ht_num is not None:
                            proportional_ht = journal_ht_num * proportion
                            proportional_journal_data['Montant du document HT'] = f"{proportional_ht:.2f}".replace('.', ',')
                        
                        if journal_marge_num is not None:
                            proportional_marge = journal_marge_num * proportion
                            proportional_journal_data['Montant marge HT'] = f"{proportional_marge:.2f}".replace('.', ',')
                            
                        print(f"         - Montants répartis : TTC={proportional_ttc:.2f}, HT={proportional_ht:.2f}")
                    else:
                        # Si pas de proportion calculable, distribuer équitablement
                        equal_proportion = 1.0 / len(numbers)
                        print(f"       - {target_ref}: proportion égale = {equal_proportion:.3f}")
                        
                        if journal_ttc_num is not None:
                            equal_ttc = journal_ttc_num * equal_proportion
                            proportional_journal_data['Montant du document TTC'] = f"{equal_ttc:.2f}".replace('.', ',')
                        
                        if journal_ht_num is not None:
                            equal_ht = journal_ht_num * equal_proportion
                            proportional_journal_data['Montant du document HT'] = f"{equal_ht:.2f}".replace('.', ',')
                        
                        if journal_marge_num is not None:
                            equal_marge = journal_marge_num * equal_proportion
                            proportional_journal_data['Montant marge HT'] = f"{equal_marge:.2f}".replace('.', ',')
                    
                    # Stocker le mapping
                    journal_mapping[target_ref] = proportional_journal_data
                    print(f"       - Mapped {target_ref} -> {proportional_journal_data['Référence LMB']} (montants répartis)")
    
    # Appliquer le mapping aux commandes
    journal_data = []
    for idx, row in df_orders_copy.iterrows():
        order_ref = row['Name_normalized']
        if order_ref in journal_mapping:
            journal_data.append(journal_mapping[order_ref])
        else:
            # Créer une ligne vide avec les mêmes colonnes
            empty_row = pd.Series(index=df_journal_copy.columns, dtype=object)
            journal_data.append(empty_row)
    
    # Convertir en DataFrame
    df_journal_mapped = pd.DataFrame(journal_data, index=df_orders_copy.index)
    
    # Concaténer horizontalement
    df_merged = pd.concat([df_orders_copy, df_journal_mapped], axis=1)
    
    # Compter les correspondances
    correspondances = df_merged['Référence LMB'].notna().sum()
    total = len(df_merged)
    
    print(f"     - Correspondances trouvées : {correspondances}/{total} ({correspondances/total*100:.1f}%)")
    
    return df_merged

def fill_missing_data_indicators(df_final, df_merged_final):
    """
    Ajoute une colonne de statut simple : COMPLET ou INCOMPLET
    Laisse les cellules vides sans marqueur pour les montants principaux (HT, TVA, TTC)
    afin que le formatage conditionnel rouge s'applique.
    """    # 1. Nettoyer SEULEMENT les colonnes numériques secondaires (pas HT, TVA, TTC, ni les méthodes de paiement)
    # Les colonnes HT, TVA, TTC gardent leurs NaN pour le formatage conditionnel rouge
    # Les colonnes de méthodes de paiement gardent leurs valeurs calculées
    secondary_numeric_columns = ['reste', 'Shopify', 'Frais de commission']
    
    for col in secondary_numeric_columns:
        if col in df_final.columns:
            df_final[col] = df_final[col].fillna(0)
      # 2. Déterminer le statut : COMPLET ou INCOMPLET
    status_info = []
    for idx, row in df_final.iterrows():
        # Une ligne est COMPLÈTE si elle a :
        # - Une référence LMB (pas vide/NaN)
        # - Au moins une méthode de paiement avec un montant > 0
        # - Une date de facture (pas vide/NaN)
        
        has_lmb = pd.notna(row['Réf. LMB']) and str(row['Réf. LMB']).strip() != ''
        has_date = pd.notna(row['Date Facture']) and str(row['Date Facture']).strip() != ''
        
        # Vérifier toutes les méthodes de paiement
        payment_methods = ['Virement bancaire', 'ALMA', 'Younited', 'PayPal', 'Shopify']
        has_payment = False
        
        for method in payment_methods:
            if method in row and pd.notna(row[method]):
                try:
                    amount = float(row[method])
                    if abs(amount) > 0.01:  # Tolérance pour les erreurs d'arrondi
                        has_payment = True
                        break
                except (ValueError, TypeError):
                    continue
        
        if has_lmb and has_payment and has_date:
            status_info.append("COMPLET")
        else:
            status_info.append("INCOMPLET")
      # 3. Préparer la colonne de statut pour les formules Excel dynamiques
    df_final['Statut'] = ''  # Colonne vide - les formules seront ajoutées dans Excel
    print(f"DEBUG: Cellules NaN conservées pour formatage rouge - HT, TVA, TTC")
    print(f"DEBUG: Cellules conservées (valeurs calculées) - Virement bancaire, ALMA, Younited, PayPal")
    print(f"DEBUG: Cellules nettoyées (NaN->0) - colonnes secondaires: {secondary_numeric_columns}")
    
    return df_final

def combine_with_old_file(df_new_data, old_file_path):
    """
    Combine les nouvelles données avec un ancien fichier Excel/CSV
    Gestion intelligente des conflits :
    - Priorité aux données de l'ancien fichier
    - Exception : nouvelles données utilisées si elles complètent des données manquantes
    """
    try:
        print("=== DÉBUT COMBINAISON INTELLIGENTE AVEC ANCIEN FICHIER ===")
        
        # Charger l'ancien fichier
        if old_file_path.endswith('.xlsx'):
            df_old = pd.read_excel(old_file_path)
            print(f"Ancien fichier Excel chargé: {len(df_old)} lignes")
        else:
            df_old = pd.read_csv(old_file_path)
            print(f"Ancien fichier CSV chargé: {len(df_old)} lignes")
        
        # Vérifier que la colonne Réf.WEB existe dans les deux fichiers
        if 'Réf.WEB' not in df_old.columns:
            print("ERREUR: La colonne 'Réf.WEB' n'existe pas dans l'ancien fichier")
            return df_new_data
        
        if 'Réf.WEB' not in df_new_data.columns:
            print("ERREUR: La colonne 'Réf.WEB' n'existe pas dans les nouvelles données")
            return df_old
        
        print(f"Nouvelles données: {len(df_new_data)} lignes")
        print(f"Colonnes anciennes: {list(df_old.columns)}")
        print(f"Colonnes nouvelles: {list(df_new_data.columns)}")
        
        # Identifier les références communes (potentiels conflits)
        old_refs = set(df_old['Réf.WEB'].dropna())
        new_refs = set(df_new_data['Réf.WEB'].dropna())
        conflicting_refs = old_refs.intersection(new_refs)
        
        print(f"Références avec conflits potentiels: {len(conflicting_refs)}")
        if conflicting_refs:
            print(f"Exemples de conflits: {list(conflicting_refs)[:5]}")
        
        # Harmoniser les colonnes d'abord
        old_columns = set(df_old.columns)
        new_columns = set(df_new_data.columns)
        all_columns = old_columns.union(new_columns)
        
        # Ajouter les colonnes manquantes avec des valeurs vides/NaN
        for col in new_columns - old_columns:
            df_old[col] = pd.NA
            print(f"Colonne '{col}' ajoutée à l'ancien fichier")
        
        for col in old_columns - new_columns:
            df_new_data[col] = pd.NA
            print(f"Colonne '{col}' ajoutée aux nouvelles données")
        
        # Réordonner les colonnes
        common_columns = sorted(all_columns)
        df_old = df_old[common_columns]
        df_new_data = df_new_data[common_columns]
        
        # Séparer les nouvelles données en deux groupes
        df_new_unique = df_new_data[~df_new_data['Réf.WEB'].isin(conflicting_refs)].copy()
        df_new_conflicts = df_new_data[df_new_data['Réf.WEB'].isin(conflicting_refs)].copy()
        
        print(f"Nouvelles données uniques (pas de conflit): {len(df_new_unique)} lignes")
        print(f"Nouvelles données en conflit: {len(df_new_conflicts)} lignes")
        
        # Traiter les conflits avec priorité intelligente
        conflicts_resolved = 0
        data_completed = 0
        
        if len(df_new_conflicts) > 0:
            print("=== RÉSOLUTION DES CONFLITS ===")
            
            for ref in conflicting_refs:
                old_row = df_old[df_old['Réf.WEB'] == ref].iloc[0]
                new_row = df_new_conflicts[df_new_conflicts['Réf.WEB'] == ref].iloc[0]
                
                # Analyser chaque colonne pour détecter les données manquantes
                for col in common_columns:
                    if col != 'Réf.WEB':  # Ne pas modifier la référence
                        old_value = old_row[col]
                        new_value = new_row[col]
                          # Vérifier si l'ancienne valeur est manquante/vide
                        old_is_empty = (pd.isna(old_value) or 
                                      old_value == '' or 
                                      old_value == 0 or 
                                      str(old_value).strip() == '' or
                                      str(old_value).lower() in ['nan', 'null', 'none', '<na>'])
                        
                        # Vérifier si la nouvelle valeur apporte des données
                        new_has_data = (not pd.isna(new_value) and 
                                      new_value != '' and 
                                      new_value != 0 and 
                                      str(new_value).strip() != '' and
                                      str(new_value).lower() not in ['nan', 'null', 'none', '<na>'])
                        
                        # Si l'ancien est vide et le nouveau a des données, on complète
                        if old_is_empty and new_has_data:
                            df_old.loc[df_old['Réf.WEB'] == ref, col] = new_value
                            data_completed += 1
                            print(f"  ✓ {ref} - Colonne '{col}': Complété '{old_value}' → '{new_value}'")
                        elif not old_is_empty and new_has_data and old_value != new_value:
                            # Conflit réel : priorité à l'ancien fichier
                            print(f"  → {ref} - Colonne '{col}': Ancien conservé '{old_value}' (nouveau: '{new_value}')")
                            conflicts_resolved += 1
                
        # Combiner : ancien fichier (mis à jour) + nouvelles données uniques
        df_combined = pd.concat([df_old, df_new_unique], ignore_index=True)
        
        print(f"=== RÉSULTAT COMBINAISON INTELLIGENTE ===")
        print(f"Total lignes combinées: {len(df_combined)}")
        print(f"Anciennes données (conservées): {len(df_old)}")
        print(f"Nouvelles données uniques ajoutées: {len(df_new_unique)}")
        print(f"Conflits résolus (priorité ancien): {conflicts_resolved}")
        print(f"Données complétées (ancien vide): {data_completed}")
        print(f"Doublons évités: {len(conflicting_refs)}")
        
        return df_combined
        
    except Exception as e:
        print(f"ERREUR lors de la combinaison intelligente: {e}")
        print("Retour des nouvelles données uniquement")
        return df_new_data

@app.route('/')
def index():
    """Page d'accueil avec le formulaire de téléchargement"""
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_files():
    """Traite les fichiers uploadés et génère le tableau consolidé"""
    print(f"\n{'*'*80}")
    print(f"ROUTE /process APPELÉE!")
    print(f"Method: {request.method}")
    print(f"Content-Type: {request.content_type}")
    print(f"Content-Length: {request.content_length}")
    print(f"Form keys: {list(request.form.keys())}")
    print(f"Files keys: {list(request.files.keys())}")
    print(f"{'*'*80}\n")
    
    logger.info("=== DEBUT DU TRAITEMENT ===")
    logger.info(f"Méthode HTTP: {request.method}")
    logger.info(f"URL: {request.url}")
    logger.info(f"Headers: {dict(request.headers)}")
    logger.info(f"Form keys: {list(request.form.keys())}")
    logger.info(f"Files keys: {list(request.files.keys())}")
    
    try:
        # Récupérer le mode de traitement
        processing_mode = request.form.get('processing_mode', 'new')
        logger.info(f"Mode de traitement: {processing_mode}")
        
        # Vérification de la présence de tous les fichiers requis
        required_files = ['orders_file', 'transactions_file', 'journal_file']
        files = {}
        
        logger.info("Vérification des fichiers requis...")
        for file_key in required_files:
            logger.debug(f"Vérification du fichier: {file_key}")
            if file_key not in request.files:
                error_msg = f'Le fichier {file_key.replace("_", " ")} est manquant.'
                logger.error(error_msg)
                flash(error_msg)
                return redirect(url_for('index'))
            
            file = request.files[file_key]
            logger.debug(f"Fichier {file_key}: nom='{file.filename}', taille={file.content_length if hasattr(file, 'content_length') else 'inconnue'}")
            
            if file.filename == '':
                error_msg = f'Veuillez sélectionner un fichier pour {file_key.replace("_", " ")}.'
                logger.error(error_msg)
                flash(error_msg)
                return redirect(url_for('index'))
            
            if not allowed_file(file.filename):
                error_msg = f'Le fichier {file.filename} doit être un fichier CSV.'
                logger.error(error_msg)
                flash(error_msg)
                return redirect(url_for('index'))
            
            files[file_key] = file
        
        logger.info("Tous les fichiers requis sont présents et valides")
        
        # En mode combinaison, vérifier la présence du fichier ancien
        old_file = None
        if processing_mode == 'combine':
            logger.info("Mode combinaison: vérification du fichier ancien...")
            if 'old_file' not in request.files or request.files['old_file'].filename == '':
                error_msg = 'Veuillez sélectionner un ancien fichier à compléter.'
                logger.error(error_msg)
                flash(error_msg)
                return redirect(url_for('index'))
            
            old_file = request.files['old_file']
            logger.debug(f"Fichier ancien: nom='{old_file.filename}'")
            
            if not (old_file.filename.endswith('.xlsx') or old_file.filename.endswith('.csv')):
                error_msg = 'L\'ancien fichier doit être au format Excel (.xlsx) ou CSV (.csv).'
                logger.error(error_msg)
                flash(error_msg)
                return redirect(url_for('index'))
            
            files['old_file'] = old_file
        
        # Sauvegarde temporaire des fichiers
        temp_paths = {}
        logger.info("Sauvegarde temporaire des fichiers...")
        try:
            for file_key, file in files.items():
                filename = secure_filename(file.filename)
                temp_path = os.path.join(UPLOAD_FOLDER, filename)
                logger.debug(f"Sauvegarde {file_key} vers: {temp_path}")
                file.save(temp_path)
                temp_paths[file_key] = temp_path
                logger.debug(f"Fichier {file_key} sauvegardé avec succès")
                
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde des fichiers: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            flash(f"Erreur lors de la sauvegarde des fichiers: {str(e)}")
            return redirect(url_for('index'))
            
        # Génération des nouvelles données
        df_new_data = generate_consolidated_billing_table(
            temp_paths['orders_file'],
            temp_paths['transactions_file'], 
            temp_paths['journal_file']
        )
        
        # Traitement selon le mode
        if processing_mode == 'combine' and old_file:
            # Mode combinaison : fusionner avec l'ancien fichier
            old_file_path = os.path.join(UPLOAD_FOLDER, secure_filename(old_file.filename))
            old_file.save(old_file_path)
            temp_paths['old_file'] = old_file_path
            
            df_result = combine_with_old_file(df_new_data, old_file_path)
            flash(f'Fichiers combinés avec succès! Fusion intelligente effectuée. Total: {len(df_result)} lignes.', 'success')
        else:
            # Mode nouveau fichier
            df_result = df_new_data
            flash(f'Tableau généré avec succès! {len(df_result)} lignes traitées.', 'success')
          # Création du fichier de sortie avec timestamp au format DD_MM_YYYY
            timestamp = datetime.now().strftime('%d_%m_%Y')
            output_filename = f'Compta_LCDI_Shopify_{timestamp}.csv'
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
              # Sauvegarde avec formatage conditionnel (Excel) ou CSV si pas possible
            final_path, is_excel = save_with_conditional_formatting(df_result, output_path)
            
            if is_excel:
                flash(f'Tableau Excel généré avec succès! {len(df_result)} lignes traitées. Les informations manquantes sont en rouge clair.', 'success')
                download_filename = os.path.basename(final_path)
            else:
                flash(f'Tableau CSV généré avec succès! {len(df_result)} lignes traitées.', 'success')
                download_filename = os.path.basename(final_path)
              # Rediriger vers la page de succès qui gère le téléchargement automatique
            return redirect(url_for('success_page', filename=download_filename))
        
    except Exception as e:
        logger.error(f'Erreur lors du traitement: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        flash(f'Erreur lors du traitement: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.errorhandler(413)
def too_large(e):
    """Gestion des fichiers trop volumineux"""
    flash('Le fichier est trop volumineux. Taille maximale: 16MB.', 'error')
    return redirect(url_for('index'))

def save_with_conditional_formatting(df_result, output_path):
    """
    Sauvegarde le DataFrame en Excel avec formatage conditionnel rouge clair 
    pour les cellules vides/manquantes
    """
    try:
        # Essayer d'importer openpyxl pour Excel
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Tableau Facturation"
          # Ajouter les données du DataFrame
        for r in dataframe_to_rows(df_result, index=False, header=True):
            ws.append(r)        # Définir les styles de formatage
        # Rouge encore plus profond pour les cellules manquantes et INCOMPLET
        missing_fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')  # Rouge encore plus profond
        incomplete_fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')  # Même rouge encore plus profond        # Vert plus profond pour COMPLET
        complete_fill = PatternFill(start_color='66CC66', end_color='66CC66', fill_type='solid')  # Vert plus profond
          # Style pour les en-têtes (gras)
        from openpyxl.styles import Font
        arial_font = Font(name='Arial', size=10)  # Police Arial par défaut
        header_font = Font(name='Arial', bold=True, size=10)
        
        # Style pour la colonne Shopify (texte rouge)
        shopify_font = Font(name='Arial', color='FF0000', bold=True, size=10)  # Rouge pour l'en-tête
        shopify_content_font = Font(name='Arial', color='FF0000', size=10)  # Rouge pour le contenu
        
        # Colonnes où vérifier les données manquantes (exclure les colonnes numériques qui sont à 0)
        important_columns = ['Réf. LMB', 'Date Facture', 'Etat', 'Client']
        header_row = [cell.value for cell in ws[1]]
        
        # Appliquer le formatage gras aux en-têtes (première ligne)
        for col_idx, cell in enumerate(ws[1]):
            if col_idx < len(header_row) and header_row[col_idx] == 'Shopify':
                # En-tête Shopify en rouge gras
                cell.font = shopify_font
            else:
                # Autres en-têtes en gras normal
                cell.font = header_font
        
        # Trouver l'index de la colonne Statut et Shopify
        statut_col_idx = None
        shopify_col_idx = None
        for idx, col_name in enumerate(header_row):
            if col_name == 'Statut':
                statut_col_idx = idx
            elif col_name == 'Shopify':
                shopify_col_idx = idx        # Appliquer le formatage aux cellules
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
            for col_idx, cell in enumerate(row):
                # Appliquer la police Arial par défaut à toutes les cellules
                cell.font = arial_font
                
                # 1. Formatage des cellules vides dans les colonnes importantes
                if col_idx < len(header_row) and header_row[col_idx] in important_columns:
                    # Vérifier si la cellule correspondante dans le DataFrame est vide/NaN
                    df_value = df_result.iloc[row_idx, col_idx]
                    
                    # Appliquer le formatage si la valeur est NaN, None, ou chaîne vide
                    if pd.isna(df_value) or df_value == '' or df_value is None:
                        cell.fill = missing_fill
                        # Laisser la cellule vide (pas de texte)
                        cell.value = None                # 2. Formatage et formules dynamiques de la colonne Statut
                elif col_idx == statut_col_idx and statut_col_idx is not None:
                    # Calculer la ligne Excel (en commençant à 2 car ligne 1 = en-têtes)
                    excel_row = row_idx + 2
                    
                    # Identifier les colonnes nécessaires pour la formule
                    ref_lmb_col = None
                    reste_col = None
                    
                    # Fonction pour convertir index en lettre(s) Excel
                    def col_index_to_letter(index):
                        """Convertit un index de colonne (0-based) en lettre(s) Excel"""
                        letter = ""
                        while index >= 0:
                            letter = chr(index % 26 + 65) + letter
                            index = index // 26 - 1
                        return letter
                    
                    for i, col_name in enumerate(header_row):
                        if col_name == 'Réf. LMB':
                            ref_lmb_col = col_index_to_letter(i)
                        elif col_name == 'reste':
                            reste_col = col_index_to_letter(i)
                    
                    # Créer la formule Excel dynamique
                    if ref_lmb_col and reste_col:
                        # Formule: SI ET les conditions sont remplies, alors "COMPLET", sinon "INCOMPLET"
                        # Conditions: Réf. LMB non vide ET reste = 0
                        formula = f'=IF(AND({ref_lmb_col}{excel_row}<>"",{reste_col}{excel_row}=0),"COMPLET","INCOMPLET")'
                        cell.value = formula
                        print(f"DEBUG: Formule ajoutée ligne {excel_row}: {formula}")
                        
                        # Appliquer un formatage neutre car le formatage conditionnel sera géré par Excel
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanc
                    else:
                        # Fallback si colonnes non trouvées
                        cell.value = "INCOMPLET"
                        cell.fill = incomplete_fill
                        print(f"DEBUG: Colonnes non trouvées pour formule, fallback ligne {excel_row}")
                
                # 3. Formatage de la colonne Shopify (texte rouge)
                elif col_idx == shopify_col_idx and shopify_col_idx is not None:
                    if cell.value is not None and cell.value != 0:
                        cell.font = shopify_content_font
          # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Appliquer le formatage conditionnel pour la colonne Statut
        if statut_col_idx is not None:
            from openpyxl.formatting.rule import CellIsRule
            
            # Convertir l'index de colonne en lettre Excel
            def col_index_to_letter(index):
                letter = ""
                while index >= 0:
                    letter = chr(index % 26 + 65) + letter
                    index = index // 26 - 1
                return letter
            
            statut_col_letter = col_index_to_letter(statut_col_idx)
            
            # Définir la plage de la colonne Statut (de ligne 2 à la dernière ligne)
            statut_range = f"{statut_col_letter}2:{statut_col_letter}{ws.max_row}"
            
            # Règle pour "COMPLET" - fond vert
            rule_complet = CellIsRule(operator='equal', formula=['"COMPLET"'], fill=complete_fill)
            
            # Règle pour "INCOMPLET" - fond rouge
            rule_incomplet = CellIsRule(operator='equal', formula=['"INCOMPLET"'], fill=incomplete_fill)
            
            # Appliquer les règles de formatage conditionnel
            ws.conditional_formatting.add(statut_range, rule_complet)
            ws.conditional_formatting.add(statut_range, rule_incomplet)
            
            print(f"DEBUG: Formatage conditionnel appliqué à la plage {statut_range}")
        
        # Figer la première ligne (en-têtes de colonnes) pour qu'elle reste visible lors du défilement
        ws.freeze_panes = 'A2'  # Fige tout ce qui est au-dessus de la ligne 2 (donc la ligne 1 avec les en-têtes)
        
        # Changer l'extension pour Excel
        excel_path = output_path.replace('.csv', '.xlsx')
        
        # Sauvegarder le fichier Excel
        wb.save(excel_path)
        
        return excel_path, True
        
    except ImportError:
        # Si openpyxl n'est pas disponible, sauvegarder en CSV normal
        print("⚠️ openpyxl non disponible, sauvegarde en CSV")
        df_result.to_csv(output_path, sep=';', decimal=',', index=False, encoding='utf-8-sig')
        return output_path, False
    except Exception as e:
        # En cas d'erreur avec Excel, fallback vers CSV
        print(f"⚠️ Erreur lors de la création Excel : {e}")
        df_result.to_csv(output_path, sep=';', decimal=',', index=False, encoding='utf-8-sig')
        return output_path, False

@app.route('/success/<filename>')
def success_page(filename):
    """Page de succès qui gère le téléchargement automatique et le rechargement"""
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    
    if os.path.exists(file_path):
        # Déterminer le type MIME
        if filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mimetype = 'text/csv'
        
        return render_template('success.html', 
                             filename=filename, 
                             file_path=url_for('download_file', filename=filename),
                             mimetype=mimetype)
    else:
        flash('Fichier non trouvé.', 'error')
        return redirect(url_for('index'))

@app.route('/download/<filename>')
def download_file(filename):
    """Route pour télécharger le fichier généré"""
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    
    if os.path.exists(file_path):
        # Déterminer le type MIME
        if filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mimetype = 'text/csv'
        
        return send_file(
            file_path,            as_attachment=True, 
            download_name=filename,
            mimetype=mimetype
        )
    else:
        flash('Fichier non trouvé.', 'error')
        return redirect(url_for('index'))

@app.route('/test-post', methods=['POST', 'GET'])
def test_post():
    """Route de test pour vérifier le fonctionnement des POST"""
    print(f"\n{'='*60}")
    print(f"ROUTE TEST-POST APPELÉE!")
    print(f"Method: {request.method}")
    print(f"{'='*60}\n")
    
    if request.method == 'POST':
        return "POST fonctionne!", 200
    else:
        return '''
        <form method="POST" action="/test-post">
            <input type="submit" value="Test POST">
        </form>
        '''

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    
    print("=== DÉMARRAGE DE L'APPLICATION ===")
    print("Application de génération de tableau de facturation LCDI")
    print(f"Port: {port}")
    print(f"Mode debug: {debug_mode}")
    print("======================================")
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
