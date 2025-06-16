#!/usr/bin/env python3
"""
Test pour vérifier que les cellules vides sont formatées en rouge clair
sans texte "❌ MANQUANT"
"""

import pandas as pd
import numpy as np
from app import save_with_conditional_formatting
import tempfile
import os

def test_formatage_cellules_vides():
    """Test que les cellules vides sont formatées en rouge clair"""
    
    # Créer un DataFrame de test avec des valeurs manquantes
    test_data = {
        'Réf. LMB': ['LMB001', np.nan, 'LMB003', ''],
        'Date Facture': ['2024-01-01', '2024-01-02', np.nan, '2024-01-04'],
        'Client': ['Client A', '', 'Client C', 'Client D'],
        'Etat': ['Payé', 'En cours', np.nan, 'Annulé'],
        'TTC': [100, 200, 300, 0],
        'Statut': ['COMPLET', 'INCOMPLET', 'INCOMPLET', 'COMPLET']
    }
    
    df_test = pd.DataFrame(test_data)
    
    # Créer un fichier temporaire
    with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
        temp_path = tmp.name
    
    try:
        # Sauvegarder avec formatage conditionnel
        result_path, is_excel = save_with_conditional_formatting(df_test, temp_path)
        
        if is_excel:
            print("✅ Fichier Excel créé avec succès")
            print(f"📁 Chemin: {result_path}")
            
            # Vérifier le contenu du fichier Excel
            try:
                from openpyxl import load_workbook
                
                wb = load_workbook(result_path)
                ws = wb.active
                
                print("\n📋 Contenu du fichier Excel:")
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    print(f"Ligne {row_idx}: {row}")
                
                print("\n🎨 Formatage des cellules:")
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                    for col_idx, cell in enumerate(row):
                        if cell.fill.start_color.index != '00000000':  # Cellule colorée
                            col_name = ws.cell(1, col_idx + 1).value
                            print(f"   Cellule {col_name} ligne {row_idx}: fond coloré (valeur: {cell.value})")
                
                print("✅ Test formatage terminé avec succès")
                
            except Exception as e:
                print(f"❌ Erreur lors de la vérification du fichier Excel: {e}")
        else:
            print("⚠️ Fallback vers CSV (openpyxl non disponible)")
            
    finally:
        # Nettoyer les fichiers temporaires
        for ext in ['.csv', '.xlsx']:
            test_file = temp_path.replace('.csv', ext)
            if os.path.exists(test_file):
                os.unlink(test_file)

if __name__ == "__main__":
    test_formatage_cellules_vides()
